from __future__ import annotations

from datetime import datetime, timedelta
import tkinter as tk
from tkinter import filedialog, messagebox, ttk


class TimeEntriesRuntimeMixin:
    """Time-entry CRUD, editing dialog, export, and invoice selection runtime behavior."""

    def edit_time_entry(self):
        selection = self.entries_tree.selection()
        if not selection:
            messagebox.showerror("Error", "Please select a time entry to edit")
            return

        selected_item = selection[0]
        item_tags = self.entries_tree.item(selected_item)["tags"]

        if "entry" not in item_tags:
            messagebox.showerror("Error", "Please select an individual time entry (not a group)")
            return

        entry_id = None
        for tag in item_tags:
            if tag.startswith("entry_id_"):
                entry_id = int(tag.replace("entry_id_", ""))
                break

        if not entry_id:
            messagebox.showerror("Error", "Could not find entry ID")
            return

        self.open_edit_time_entry_dialog(entry_id)

    def open_edit_time_entry_dialog_and_refresh(self, entry_id):
        """Open edit dialog and refresh invoice list after closing."""
        edit_window = tk.Toplevel(self.root)
        edit_window.title("Edit Time Entry")
        self.center_dialog(edit_window, 500, 400)

        edit_window.transient(self.root)
        edit_window.grab_set()

        edit_window.update_idletasks()
        x = self.root.winfo_x() + (self.root.winfo_width() // 2) - (edit_window.winfo_width() // 2)
        y = self.root.winfo_y() + (self.root.winfo_height() // 2) - (edit_window.winfo_height() // 2)
        edit_window.geometry(f"+{x}+{y}")

        edit_window.lift()
        edit_window.focus_force()

        def on_close():
            edit_window.destroy()
            if hasattr(self, "invoice_entries_tree"):
                self.load_invoiceable_entries()

        edit_window.protocol("WM_DELETE_WINDOW", on_close)
        self._build_edit_dialog_content(edit_window, entry_id, on_close)

    def open_edit_time_entry_dialog(self, entry_id):
        """Original edit dialog (used by Time Entries tab)."""
        edit_window = tk.Toplevel(self.root)
        edit_window.title("Edit Time Entry")
        self.center_dialog(edit_window, 500, 400)

        edit_window.transient(self.root)

        edit_window.update_idletasks()
        x = self.root.winfo_x() + (self.root.winfo_width() // 2) - (edit_window.winfo_width() // 2)
        y = self.root.winfo_y() + (self.root.winfo_height() // 2) - (edit_window.winfo_height() // 2)
        edit_window.geometry(f"+{x}+{y}")

        edit_window.lift()
        edit_window.focus_force()
        self._build_edit_dialog_content(edit_window, entry_id, lambda: edit_window.destroy())

    def _build_edit_dialog_content(self, edit_window, entry_id, close_callback):
        with self.db.get_connection() as conn:
            cursor = conn.cursor()
            cursor.execute("SELECT * FROM time_entries WHERE id = ?", (entry_id,))
            entry = cursor.fetchone()

        if not entry:
            messagebox.showerror("Error", f"Time entry #{entry_id} not found in database")
            close_callback()
            return

        form_frame = ttk.Frame(edit_window)
        form_frame.pack(fill="both", expand=True, padx=15, pady=15)

        task_name = entry[2] if entry[2] else "Unknown"
        project_name = entry[4] if entry[4] else "Unknown"
        client_name = entry[6] if entry[6] else "Unknown"

        ttk.Label(
            form_frame,
            text=f"Task: {client_name} - {project_name} - {task_name}",
            font=("Arial", 9, "bold"),
        ).grid(row=0, column=0, columnspan=2, sticky="w", pady=(0, 10))

        try:
            original_start = datetime.fromisoformat(entry[8])
            original_end = datetime.fromisoformat(entry[9])
            original_duration = (original_end - original_start).total_seconds() / 3600
        except Exception:
            original_start = datetime.now()
            original_end = datetime.now()
            original_duration = 0

        ttk.Label(form_frame, text="Date (MM/DD/YY):").grid(row=1, column=0, sticky="w", pady=5)
        date_entry = ttk.Entry(form_frame, width=15)
        date_entry.grid(row=1, column=1, sticky="w", padx=5, pady=5)
        date_entry.insert(0, original_start.strftime("%m/%d/%y"))

        ttk.Label(form_frame, text="Entry Mode:").grid(row=2, column=0, sticky="w", pady=5)
        mode_frame = ttk.Frame(form_frame)
        mode_frame.grid(row=2, column=1, sticky="w", padx=5, pady=5)

        edit_entry_mode = tk.StringVar(value="decimal")
        ttk.Radiobutton(mode_frame, text="Start/End Time", variable=edit_entry_mode, value="time_range").pack(
            side="left"
        )
        ttk.Radiobutton(mode_frame, text="Decimal Hours", variable=edit_entry_mode, value="decimal").pack(
            side="left", padx=10
        )

        start_time_label = ttk.Label(form_frame, text="Start Time (HH:MM AM/PM):")
        start_time_label.grid(row=3, column=0, sticky="w", pady=5)
        start_time_entry = ttk.Entry(form_frame, width=20)
        start_time_entry.grid(row=3, column=1, sticky="w", padx=5, pady=5)
        start_time_entry.insert(0, original_start.strftime("%I:%M %p"))

        end_time_label = ttk.Label(form_frame, text="End Time (HH:MM AM/PM):")
        end_time_label.grid(row=4, column=0, sticky="w", pady=5)
        end_time_entry = ttk.Entry(form_frame, width=20)
        end_time_entry.grid(row=4, column=1, sticky="w", padx=5, pady=5)
        end_time_entry.insert(0, original_end.strftime("%I:%M %p"))

        hours_label = ttk.Label(form_frame, text="Duration (hours):")
        hours_label.grid(row=5, column=0, sticky="w", pady=5)
        hours_entry = ttk.Entry(form_frame, width=15)
        hours_entry.grid(row=5, column=1, sticky="w", padx=5, pady=5)
        hours_entry.insert(0, f"{original_duration:.2f}")

        hours_help = ttk.Label(form_frame, text="(e.g., 1.5, 2.25, 0.75)", font=("Arial", 8), foreground="gray")
        hours_help.grid(row=6, column=1, sticky="w", padx=5)

        def toggle_edit_mode():
            mode = edit_entry_mode.get()
            if mode == "time_range":
                start_time_label.grid()
                start_time_entry.grid()
                end_time_label.grid()
                end_time_entry.grid()
                hours_label.grid_remove()
                hours_entry.grid_remove()
                hours_help.grid_remove()
            else:
                start_time_label.grid_remove()
                start_time_entry.grid_remove()
                end_time_label.grid_remove()
                end_time_entry.grid_remove()
                hours_label.grid()
                hours_entry.grid()
                hours_help.grid()

        edit_entry_mode.trace("w", lambda *args: toggle_edit_mode())
        toggle_edit_mode()

        ttk.Label(form_frame, text="Description:").grid(row=7, column=0, sticky="nw", pady=5)
        desc_text = tk.Text(form_frame, width=40, height=4)
        desc_text.grid(row=7, column=1, sticky="ew", padx=5, pady=5)
        desc_text.insert("1.0", entry[11] or "")

        form_frame.columnconfigure(1, weight=1)

        button_frame = ttk.Frame(edit_window)
        button_frame.pack(fill="x", padx=15, pady=(0, 15))

        def save_changes():
            try:
                date_str = date_entry.get().strip()
                description = desc_text.get("1.0", tk.END).strip()
                mode = edit_entry_mode.get()

                if mode == "time_range":
                    start_str = start_time_entry.get().strip()
                    end_str = end_time_entry.get().strip()

                    start_time = datetime.strptime(f"{date_str} {start_str}", "%m/%d/%y %I:%M %p")
                    end_time = datetime.strptime(f"{date_str} {end_str}", "%m/%d/%y %I:%M %p")

                    if end_time <= start_time:
                        messagebox.showerror("Error", "End time must be after start time")
                        return
                else:
                    hours_str = hours_entry.get().strip()
                    if not hours_str:
                        messagebox.showerror("Error", "Please enter duration in hours")
                        return

                    try:
                        duration_hours = float(hours_str)
                    except ValueError:
                        messagebox.showerror("Error", "Invalid duration format. Use decimal (e.g., 1.5)")
                        return

                    if duration_hours <= 0:
                        messagebox.showerror("Error", "Duration must be greater than 0")
                        return

                    if duration_hours > 24:
                        messagebox.showerror("Error", "Duration cannot exceed 24 hours")
                        return

                    start_time = datetime.strptime(f"{date_str} 09:00 AM", "%m/%d/%y %I:%M %p")
                    end_time = start_time + timedelta(hours=duration_hours)

                duration_hours = (end_time - start_time).total_seconds() / 3600

                self.time_entry_model.update(entry_id, start_time, end_time, description)
                self.refresh_time_entries()
                close_callback()
                messagebox.showinfo(
                    "Success",
                    f"Time entry updated successfully\n\n"
                    f"Duration: {duration_hours:.2f} hours\n\n"
                    "Click REFRESH to update the invoice data below.",
                )

            except ValueError as e:
                messagebox.showerror("Error", f"Invalid input: {str(e)}")

        ttk.Button(button_frame, text="Save Changes", command=save_changes).pack(side="right", padx=5)
        ttk.Button(button_frame, text="Cancel", command=close_callback).pack(side="right")

    def delete_time_entry(self):
        selection = self.entries_tree.selection()
        if not selection:
            messagebox.showerror("Error", "Please select a time entry to delete")
            return

        selected_item = selection[0]
        item_tags = self.entries_tree.item(selected_item)["tags"]

        if "entry" not in item_tags:
            messagebox.showerror("Error", "Please select an individual time entry (not a group)")
            return

        entry_id = None
        for tag in item_tags:
            if tag.startswith("entry_id_"):
                entry_id = int(tag.replace("entry_id_", ""))
                break

        if not entry_id:
            messagebox.showerror("Error", "Could not find entry ID")
            return

        if messagebox.askyesno("Confirm", "Delete this time entry?"):
            self.time_entry_model.delete(entry_id)
            self.refresh_time_entries()
            messagebox.showinfo("Success", "Time entry deleted successfully")

    def export_time_entries_to_excel(self):
        """Export time entries to Excel file, respecting current filter or selection."""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Alignment, Font, PatternFill
        except ImportError as e:
            messagebox.showerror(
                "Error",
                f"openpyxl library not installed.\n\nError: {e}\n\nInstall with: pip install openpyxl",
            )
            return

        try:
            selected_items = self.entries_tree.selection()

            if selected_items:
                selected_ids = []
                for item in selected_items:
                    item_data = self.entries_tree.item(item)
                    tags = item_data["tags"]
                    if tags:
                        for tag in tags:
                            if tag.startswith("entry_id_"):
                                try:
                                    entry_id = int(tag.replace("entry_id_", ""))
                                    selected_ids.append(entry_id)
                                    break
                                except (ValueError, TypeError):
                                    pass

                if not selected_ids:
                    messagebox.showinfo("No Selection", "Please select time entries (not group headers) to export.")
                    return

                with self.db.get_connection() as conn:
                    cursor = conn.cursor()
                    placeholders = ",".join("?" * len(selected_ids))
                    cursor.execute(
                        f"""
                        SELECT te.id, te.client_name, te.project_name, te.task_name,
                               te.start_time, te.end_time, te.duration_minutes, te.description,
                               te.is_billed, te.invoice_number, te.date
                        FROM time_entries te
                        WHERE te.id IN ({placeholders})
                        ORDER BY te.date DESC, te.start_time DESC
                    """,
                        selected_ids,
                    )
                    entries = cursor.fetchall()

                export_scope = "Selected"
            else:
                filter_val = self.time_entries_filter_var.get() if hasattr(self, "time_entries_filter_var") else "unbilled"

                where_clause = ""
                if filter_val == "unbilled":
                    where_clause = "WHERE te.is_billed = 0"
                elif filter_val == "billed":
                    where_clause = "WHERE te.is_billed = 1"

                with self.db.get_connection() as conn:
                    cursor = conn.cursor()
                    cursor.execute(
                        f"""
                        SELECT te.id, te.client_name, te.project_name, te.task_name,
                               te.start_time, te.end_time, te.duration_minutes, te.description,
                               te.is_billed, te.invoice_number, te.date
                        FROM time_entries te
                        {where_clause}
                        ORDER BY te.date DESC, te.start_time DESC
                    """
                    )
                    entries = cursor.fetchall()

                export_scope = filter_val.capitalize()

            if not entries:
                messagebox.showinfo("No Data", "No time entries found to export.")
                return

            filename = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[("Excel files", "*.xlsx")],
                initialfile=f"TimeEntries_{export_scope}_{datetime.now().strftime('%Y%m%d')}.xlsx",
            )

            if not filename:
                return

            wb = Workbook()
            ws = wb.active
            ws.title = f"Time Entries ({export_scope})"

            headers = [
                "Date",
                "Client",
                "Project",
                "Task",
                "Start Time",
                "End Time",
                "Duration (hrs)",
                "Description",
                "Billed",
                "Invoice #",
            ]
            ws.append(headers)

            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center")

            for entry in entries:
                (
                    entry_id,
                    client,
                    project,
                    task,
                    start_time,
                    end_time,
                    duration_mins,
                    description,
                    is_billed,
                    invoice_num,
                    date,
                ) = entry

                try:
                    start_dt = datetime.fromisoformat(start_time)
                    start_display = start_dt.strftime("%I:%M %p")
                except Exception:
                    start_display = start_time

                try:
                    end_dt = datetime.fromisoformat(end_time)
                    end_display = end_dt.strftime("%I:%M %p")
                except Exception:
                    end_display = end_time

                duration_hours = (duration_mins / 60.0) if duration_mins else 0
                billed_status = "Yes" if is_billed else "No"

                ws.append(
                    [
                        date,
                        client,
                        project,
                        task,
                        start_display,
                        end_display,
                        round(duration_hours, 2),
                        description or "",
                        billed_status,
                        invoice_num or "",
                    ]
                )

            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except Exception:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width

            wb.save(filename)
            messagebox.showinfo("Success", f"Exported {len(entries)} time entries to:\n\n{filename}")

            if messagebox.askyesno("Open File?", "Would you like to open the exported file now?"):
                import os

                os.startfile(filename)

        except Exception as e:
            messagebox.showerror("Export Error", f"Failed to export time entries:\n\n{e}\n\n{type(e).__name__}")
            import traceback

            traceback.print_exc()

    def invoice_selected_entries(self):
        """Generate invoice from selected time entries."""
        selection = self.entries_tree.selection()
        if not selection:
            messagebox.showerror("Error", "Please select one or more time entries to invoice")
            return

        entry_ids = []
        billed_entries = []
        client_id = None
        client_name = None

        for item in selection:
            item_tags = self.entries_tree.item(item)["tags"]
            if "entry" not in item_tags:
                continue

            entry_id = None
            for tag in item_tags:
                if tag.startswith("entry_id_"):
                    entry_id = int(tag.replace("entry_id_", ""))
                    break

            if not entry_id:
                continue

            all_entries = self.time_entry_model.get_all()
            entry_data = None
            for e in all_entries:
                if e[0] == entry_id:
                    entry_data = e
                    break

            if not entry_data:
                continue

            entry_client = entry_data[1]
            entry_task = entry_data[3]

            if "[BILLED]" in entry_task:
                billed_entries.append(entry_id)
                continue

            entry_ids.append(entry_id)

            if client_id is None:
                client_name = entry_client
                clients = self.client_model.get_all()
                for client in clients:
                    if client[1] == entry_client:
                        client_id = client[0]
                        break

        if billed_entries:
            messagebox.showwarning(
                "Already Billed",
                f"{len(billed_entries)} selected entr"
                f"{'y is' if len(billed_entries) == 1 else 'ies are'} already billed and will be skipped.\n\n"
                "Click OK to continue with unbilled entries only.",
            )

        if not entry_ids:
            messagebox.showerror("Error", "No unbilled entries selected")
            return

        if not client_id:
            messagebox.showerror("Error", "Could not determine client for selected entries")
            return

        for item in selection:
            if self.entries_tree.item(item)["values"][1] != client_name:
                messagebox.showerror(
                    "Error",
                    "All selected entries must be from the same client.\nPlease select entries from one client only.",
                )
                return

        if not messagebox.askyesno(
            "Confirm Invoice",
            f"Generate invoice for {len(entry_ids)} time entry(ies) from {client_name}?",
        ):
            return

        self.generate_invoice_from_entries(client_id, client_name, entry_ids)
