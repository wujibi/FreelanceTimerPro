"""CustomTkinter Time Entries tab — grouped tree, edit/delete/export (Tk parity)."""

from __future__ import annotations

import os
from collections.abc import Callable
from datetime import datetime, timedelta
from tkinter import filedialog, messagebox, ttk
from typing import Any

import tkinter as tk

import customtkinter as ctk

from models import TimeEntry
from ui_helpers import center_dialog, restore_tree_state, save_tree_state


class CtkTimeEntriesTab:
    def __init__(self, parent: Any, root: Any, db) -> None:
        self.parent = parent
        self.root = root
        self.db = db
        self.time_entry_model = TimeEntry(self.db)

        self.time_entries_filter_var = tk.StringVar(value="unbilled")

        self._build_ui()
        self.refresh()

    def _build_ui(self) -> None:
        head = ctk.CTkLabel(
            self.parent,
            text="Time Entries (grouped by Client → Project → Task)",
            font=ctk.CTkFont(size=14, weight="bold"),
        )
        head.pack(anchor="w", padx=10, pady=(8, 4))

        filt = ctk.CTkFrame(self.parent, fg_color="transparent")
        filt.pack(fill="x", padx=10, pady=4)
        ctk.CTkLabel(filt, text="Show:", font=ctk.CTkFont(weight="bold")).pack(side="left", padx=(0, 8))
        for text, val in (
            ("Unbilled only", "unbilled"),
            ("Billed only", "billed"),
            ("All entries", "all"),
        ):
            ctk.CTkRadioButton(
                filt,
                text=text,
                variable=self.time_entries_filter_var,
                value=val,
                command=self.refresh,
            ).pack(side="left", padx=6)

        ctk.CTkLabel(
            self.parent,
            text="Tip: expand ▶ groups, select a row labeled “Entry”, then Edit or Delete.",
            text_color=("gray30", "gray65"),
            font=ctk.CTkFont(size=11),
        ).pack(anchor="w", padx=10, pady=(0, 6))

        btns = ctk.CTkFrame(self.parent, fg_color="transparent")
        btns.pack(fill="x", padx=10, pady=4)
        ctk.CTkButton(btns, text="Edit entry", command=self.edit_time_entry).pack(side="left", padx=4)
        ctk.CTkButton(btns, text="Delete entry", command=self.delete_time_entry).pack(side="left", padx=4)
        ctk.CTkButton(btns, text="Export to Excel", command=self.export_time_entries_to_excel).pack(
            side="left", padx=4
        )

        # ttk.Treeview must sit in a tkinter Frame (tab frame is the logical parent).
        self._tree_host = tk.Frame(self.parent)
        self._tree_host.pack(fill="both", expand=True, padx=8, pady=(4, 10))

        self.entries_tree = ttk.Treeview(
            self._tree_host,
            columns=("Type", "Name", "Start", "Duration", "Description"),
            selectmode="extended",
        )
        self.entries_tree.heading("#0", text="Hierarchy")
        self.entries_tree.heading("Type", text="Type")
        self.entries_tree.heading("Name", text="Details")
        self.entries_tree.heading("Start", text="Start Time")
        self.entries_tree.heading("Duration", text="Duration")
        self.entries_tree.heading("Description", text="Description")
        self.entries_tree.column("#0", width=240)
        self.entries_tree.column("Type", width=72)
        self.entries_tree.column("Name", width=120)
        self.entries_tree.column("Start", width=130)
        self.entries_tree.column("Duration", width=88)
        self.entries_tree.column("Description", width=220)

        scroll = ttk.Scrollbar(self._tree_host, orient="vertical", command=self.entries_tree.yview)
        self.entries_tree.configure(yscrollcommand=scroll.set)
        self.entries_tree.pack(side="left", fill="both", expand=True)
        scroll.pack(side="right", fill="y")

        self.entries_tree.tag_configure(
            "client_row",
            background="#e8f4f8",
            foreground="#13100f",
            font=("Segoe UI", 10, "bold"),
        )
        self.entries_tree.tag_configure(
            "project_row",
            background="#e8f4f8",
            foreground="#13100f",
            font=("Segoe UI", 9, "bold"),
        )
        self.entries_tree.tag_configure(
            "task_row",
            background="#e8f4f8",
            foreground="#13100f",
            font=("Segoe UI", 9),
        )
        self.entries_tree.tag_configure("entry_row", background="white", foreground="#222", font=("Segoe UI", 9))

    def refresh(self) -> None:
        expanded_items = save_tree_state(self.entries_tree)

        for item in self.entries_tree.get_children():
            self.entries_tree.delete(item)

        filter_val = self.time_entries_filter_var.get()
        where_clause = ""
        if filter_val == "unbilled":
            where_clause = "WHERE te.is_billed = 0"
        elif filter_val == "billed":
            where_clause = "WHERE te.is_billed = 1"

        with self.db.get_connection() as conn:
            cursor = conn.cursor()
            cursor.execute(
                f"""
                SELECT te.id, te.client_name, te.project_name, te.task_name,
                       te.start_time, te.end_time, te.duration_minutes, te.description,
                       te.is_billed, te.invoice_number
                FROM time_entries te
                {where_clause}
                ORDER BY te.client_name, te.project_name, te.task_name, te.start_time DESC
            """
            )
            entries = cursor.fetchall()

        if not entries:
            restore_tree_state(self.entries_tree, expanded_items, expand_all=False)
            return

        hierarchy = {}
        for entry in entries:
            client_name = entry[1]
            project_name = entry[2]
            task_name = entry[3]
            if client_name not in hierarchy:
                hierarchy[client_name] = {}
            if project_name not in hierarchy[client_name]:
                hierarchy[client_name][project_name] = {}
            if task_name not in hierarchy[client_name][project_name]:
                hierarchy[client_name][project_name][task_name] = []
            hierarchy[client_name][project_name][task_name].append(entry)

        for client_name in sorted(hierarchy.keys()):
            client_total_minutes = 0
            for project_name in hierarchy[client_name]:
                for task_name in hierarchy[client_name][project_name]:
                    for entry in hierarchy[client_name][project_name][task_name]:
                        client_total_minutes += entry[6] if entry[6] else 0

            client_total_hours = client_total_minutes / 60.0
            client_id = self.entries_tree.insert(
                "",
                "end",
                text=f"📁 {client_name}",
                values=("Client", "", "", f"{client_total_hours:.2f} hrs", ""),
                tags=("client", "client_row"),
            )

            for project_name in sorted(hierarchy[client_name].keys()):
                project_total_minutes = 0
                for task_name in hierarchy[client_name][project_name]:
                    for entry in hierarchy[client_name][project_name][task_name]:
                        project_total_minutes += entry[6] if entry[6] else 0

                project_total_hours = project_total_minutes / 60.0
                project_id = self.entries_tree.insert(
                    client_id,
                    "end",
                    text=f"  📂 {project_name}",
                    values=("Project", "", "", f"{project_total_hours:.2f} hrs", ""),
                    tags=("project", "project_row"),
                )

                for task_name in sorted(hierarchy[client_name][project_name].keys()):
                    task_entries = hierarchy[client_name][project_name][task_name]
                    task_total_minutes = sum(e[6] if e[6] else 0 for e in task_entries)
                    task_total_hours = task_total_minutes / 60.0
                    has_billed = any(e[8] for e in task_entries)
                    billed_indicator = " [BILLED]" if has_billed else ""

                    task_id = self.entries_tree.insert(
                        project_id,
                        "end",
                        text=f"    📋 {task_name}{billed_indicator}",
                        values=("Task", f"{len(task_entries)} entries", "", f"{task_total_hours:.2f} hrs", ""),
                        tags=("task", "task_row"),
                    )

                    for entry in task_entries:
                        duration_minutes = entry[6] if entry[6] else 0
                        duration_hours = duration_minutes / 60.0
                        start_time = entry[4]
                        try:
                            dt = datetime.fromisoformat(start_time)
                            start_display = dt.strftime("%m/%d/%y %I:%M %p")
                        except Exception:
                            start_display = start_time

                        entry_billed = " [BILLED]" if entry[8] else ""
                        self.entries_tree.insert(
                            task_id,
                            "end",
                            text="      ⏱️ Entry",
                            values=("Entry" + entry_billed, "", start_display, f"{duration_hours:.2f} hrs", entry[7] or ""),
                            tags=("entry", f"entry_id_{entry[0]}", "entry_row"),
                        )

        self.entries_tree.tag_configure("client", font=("Segoe UI", 10, "bold"))
        self.entries_tree.tag_configure("project", font=("Segoe UI", 9, "bold"))
        self.entries_tree.tag_configure("task", font=("Segoe UI", 9))
        self.entries_tree.tag_configure("entry", font=("Segoe UI", 8))
        restore_tree_state(self.entries_tree, expanded_items, expand_all=True)

    def edit_time_entry(self) -> None:
        selection = self.entries_tree.selection()
        if not selection:
            messagebox.showerror("Error", "Please select a time entry to edit")
            return

        selected_item = selection[0]
        item_tags = self.entries_tree.item(selected_item)["tags"]

        if "entry" not in item_tags:
            messagebox.showerror("Error", "Please select an individual time entry (not a group)")
            return

        entry_id = None
        for tag in item_tags:
            if tag.startswith("entry_id_"):
                entry_id = int(tag.replace("entry_id_", ""))
                break

        if not entry_id:
            messagebox.showerror("Error", "Could not find entry ID")
            return

        self._open_edit_time_entry_dialog(entry_id)

    def open_edit_entry_dialog(
        self,
        entry_id: int,
        after_save: Callable[[], None] | None = None,
    ) -> None:
        """Open the standard edit dialog; optional callback runs after a successful save."""
        self._open_edit_time_entry_dialog(entry_id, after_save=after_save)

    def _open_edit_time_entry_dialog(
        self,
        entry_id: int,
        after_save: Callable[[], None] | None = None,
    ) -> None:
        edit_window = tk.Toplevel(self.root)
        edit_window.title("Edit Time Entry")
        center_dialog(self.root, edit_window, 500, 400)
        edit_window.transient(self.root)
        edit_window.lift()
        edit_window.focus_force()
        self._build_edit_dialog_content(edit_window, entry_id, lambda: edit_window.destroy(), after_save=after_save)

    def _build_edit_dialog_content(
        self,
        edit_window,
        entry_id,
        close_callback,
        after_save: Callable[[], None] | None = None,
    ) -> None:
        with self.db.get_connection() as conn:
            cursor = conn.cursor()
            cursor.execute("SELECT * FROM time_entries WHERE id = ?", (entry_id,))
            entry = cursor.fetchone()

        if not entry:
            messagebox.showerror("Error", f"Time entry #{entry_id} not found in database")
            close_callback()
            return

        form_frame = ttk.Frame(edit_window)
        form_frame.pack(fill="both", expand=True, padx=15, pady=15)

        task_name = entry[2] if entry[2] else "Unknown"
        project_name = entry[4] if entry[4] else "Unknown"
        client_name = entry[6] if entry[6] else "Unknown"

        ttk.Label(
            form_frame,
            text=f"Task: {client_name} - {project_name} - {task_name}",
            font=("Arial", 9, "bold"),
        ).grid(row=0, column=0, columnspan=2, sticky="w", pady=(0, 10))

        try:
            original_start = datetime.fromisoformat(entry[8])
            original_end = datetime.fromisoformat(entry[9])
            original_duration = (original_end - original_start).total_seconds() / 3600
        except Exception:
            original_start = datetime.now()
            original_end = datetime.now()
            original_duration = 0

        ttk.Label(form_frame, text="Date (MM/DD/YY):").grid(row=1, column=0, sticky="w", pady=5)
        date_entry = ttk.Entry(form_frame, width=15)
        date_entry.grid(row=1, column=1, sticky="w", padx=5, pady=5)
        date_entry.insert(0, original_start.strftime("%m/%d/%y"))

        ttk.Label(form_frame, text="Entry Mode:").grid(row=2, column=0, sticky="w", pady=5)
        mode_frame = ttk.Frame(form_frame)
        mode_frame.grid(row=2, column=1, sticky="w", padx=5, pady=5)

        edit_entry_mode = tk.StringVar(value="decimal")
        ttk.Radiobutton(mode_frame, text="Start/End Time", variable=edit_entry_mode, value="time_range").pack(
            side="left"
        )
        ttk.Radiobutton(mode_frame, text="Decimal Hours", variable=edit_entry_mode, value="decimal").pack(
            side="left", padx=10
        )

        start_time_label = ttk.Label(form_frame, text="Start Time (HH:MM AM/PM):")
        start_time_label.grid(row=3, column=0, sticky="w", pady=5)
        start_time_entry = ttk.Entry(form_frame, width=20)
        start_time_entry.grid(row=3, column=1, sticky="w", padx=5, pady=5)
        start_time_entry.insert(0, original_start.strftime("%I:%M %p"))

        end_time_label = ttk.Label(form_frame, text="End Time (HH:MM AM/PM):")
        end_time_label.grid(row=4, column=0, sticky="w", pady=5)
        end_time_entry = ttk.Entry(form_frame, width=20)
        end_time_entry.grid(row=4, column=1, sticky="w", padx=5, pady=5)
        end_time_entry.insert(0, original_end.strftime("%I:%M %p"))

        hours_label = ttk.Label(form_frame, text="Duration (hours):")
        hours_label.grid(row=5, column=0, sticky="w", pady=5)
        hours_entry = ttk.Entry(form_frame, width=15)
        hours_entry.grid(row=5, column=1, sticky="w", padx=5, pady=5)
        hours_entry.insert(0, f"{original_duration:.2f}")

        hours_help = ttk.Label(form_frame, text="(e.g., 1.5, 2.25, 0.75)", font=("Arial", 8), foreground="gray")
        hours_help.grid(row=6, column=1, sticky="w", padx=5)

        def toggle_edit_mode():
            mode = edit_entry_mode.get()
            if mode == "time_range":
                start_time_label.grid()
                start_time_entry.grid()
                end_time_label.grid()
                end_time_entry.grid()
                hours_label.grid_remove()
                hours_entry.grid_remove()
                hours_help.grid_remove()
            else:
                start_time_label.grid_remove()
                start_time_entry.grid_remove()
                end_time_label.grid_remove()
                end_time_entry.grid_remove()
                hours_label.grid()
                hours_entry.grid()
                hours_help.grid()

        edit_entry_mode.trace("w", lambda *args: toggle_edit_mode())
        toggle_edit_mode()

        ttk.Label(form_frame, text="Description:").grid(row=7, column=0, sticky="nw", pady=5)
        desc_text = tk.Text(form_frame, width=40, height=4)
        desc_text.grid(row=7, column=1, sticky="ew", padx=5, pady=5)
        desc_text.insert("1.0", entry[11] or "")

        form_frame.columnconfigure(1, weight=1)

        button_frame = ttk.Frame(edit_window)
        button_frame.pack(fill="x", padx=15, pady=(0, 15))

        def save_changes():
            try:
                date_str = date_entry.get().strip()
                description = desc_text.get("1.0", tk.END).strip()
                mode = edit_entry_mode.get()

                if mode == "time_range":
                    start_str = start_time_entry.get().strip()
                    end_str = end_time_entry.get().strip()

                    start_time = datetime.strptime(f"{date_str} {start_str}", "%m/%d/%y %I:%M %p")
                    end_time = datetime.strptime(f"{date_str} {end_str}", "%m/%d/%y %I:%M %p")

                    if end_time <= start_time:
                        messagebox.showerror("Error", "End time must be after start time")
                        return
                else:
                    hours_str = hours_entry.get().strip()
                    if not hours_str:
                        messagebox.showerror("Error", "Please enter duration in hours")
                        return

                    try:
                        duration_hours = float(hours_str)
                    except ValueError:
                        messagebox.showerror("Error", "Invalid duration format. Use decimal (e.g., 1.5)")
                        return

                    if duration_hours <= 0:
                        messagebox.showerror("Error", "Duration must be greater than 0")
                        return

                    if duration_hours > 24:
                        messagebox.showerror("Error", "Duration cannot exceed 24 hours")
                        return

                    start_time = datetime.strptime(f"{date_str} 09:00 AM", "%m/%d/%y %I:%M %p")
                    end_time = start_time + timedelta(hours=duration_hours)

                duration_hours = (end_time - start_time).total_seconds() / 3600

                self.time_entry_model.update(entry_id, start_time, end_time, description)
                self.refresh()
                if after_save:
                    after_save()
                close_callback()
                messagebox.showinfo(
                    "Success",
                    f"Time entry updated successfully\n\n"
                    f"Duration: {duration_hours:.2f} hours\n\n"
                    "Click REFRESH to update the invoice data below.",
                )

            except ValueError as e:
                messagebox.showerror("Error", f"Invalid input: {str(e)}")

        ttk.Button(button_frame, text="Save Changes", command=save_changes).pack(side="right", padx=5)
        ttk.Button(button_frame, text="Cancel", command=close_callback).pack(side="right")

    def delete_time_entry(self) -> None:
        selection = self.entries_tree.selection()
        if not selection:
            messagebox.showerror("Error", "Please select a time entry to delete")
            return

        selected_item = selection[0]
        item_tags = self.entries_tree.item(selected_item)["tags"]

        if "entry" not in item_tags:
            messagebox.showerror("Error", "Please select an individual time entry (not a group)")
            return

        entry_id = None
        for tag in item_tags:
            if tag.startswith("entry_id_"):
                entry_id = int(tag.replace("entry_id_", ""))
                break

        if not entry_id:
            messagebox.showerror("Error", "Could not find entry ID")
            return

        if messagebox.askyesno("Confirm", "Delete this time entry?"):
            self.time_entry_model.delete(entry_id)
            self.refresh()
            messagebox.showinfo("Success", "Time entry deleted successfully")

    def export_time_entries_to_excel(self) -> None:
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Alignment, Font, PatternFill
        except ImportError as e:
            messagebox.showerror(
                "Error",
                f"openpyxl library not installed.\n\nError: {e}\n\nInstall with: pip install openpyxl",
            )
            return

        try:
            selected_items = self.entries_tree.selection()

            if selected_items:
                selected_ids = []
                for item in selected_items:
                    item_data = self.entries_tree.item(item)
                    tags = item_data["tags"]
                    if tags:
                        for tag in tags:
                            if tag.startswith("entry_id_"):
                                try:
                                    entry_id = int(tag.replace("entry_id_", ""))
                                    selected_ids.append(entry_id)
                                    break
                                except (ValueError, TypeError):
                                    pass

                if not selected_ids:
                    messagebox.showinfo("No Selection", "Please select time entries (not group headers) to export.")
                    return

                with self.db.get_connection() as conn:
                    cursor = conn.cursor()
                    placeholders = ",".join("?" * len(selected_ids))
                    cursor.execute(
                        f"""
                        SELECT te.id, te.client_name, te.project_name, te.task_name,
                               te.start_time, te.end_time, te.duration_minutes, te.description,
                               te.is_billed, te.invoice_number, te.date
                        FROM time_entries te
                        WHERE te.id IN ({placeholders})
                        ORDER BY te.date DESC, te.start_time DESC
                    """,
                        selected_ids,
                    )
                    entries = cursor.fetchall()

                export_scope = "Selected"
            else:
                filter_val = self.time_entries_filter_var.get()

                where_clause = ""
                if filter_val == "unbilled":
                    where_clause = "WHERE te.is_billed = 0"
                elif filter_val == "billed":
                    where_clause = "WHERE te.is_billed = 1"

                with self.db.get_connection() as conn:
                    cursor = conn.cursor()
                    cursor.execute(
                        f"""
                        SELECT te.id, te.client_name, te.project_name, te.task_name,
                               te.start_time, te.end_time, te.duration_minutes, te.description,
                               te.is_billed, te.invoice_number, te.date
                        FROM time_entries te
                        {where_clause}
                        ORDER BY te.date DESC, te.start_time DESC
                    """
                    )
                    entries = cursor.fetchall()

                export_scope = filter_val.capitalize()

            if not entries:
                messagebox.showinfo("No Data", "No time entries found to export.")
                return

            filename = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[("Excel files", "*.xlsx")],
                initialfile=f"TimeEntries_{export_scope}_{datetime.now().strftime('%Y%m%d')}.xlsx",
            )

            if not filename:
                return

            wb = Workbook()
            ws = wb.active
            ws.title = f"Time Entries ({export_scope})"

            headers = [
                "Date",
                "Client",
                "Project",
                "Task",
                "Start Time",
                "End Time",
                "Duration (hrs)",
                "Description",
                "Billed",
                "Invoice #",
            ]
            ws.append(headers)

            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center")

            for entry in entries:
                (
                    _entry_id,
                    client,
                    project,
                    task,
                    start_time,
                    end_time,
                    duration_mins,
                    description,
                    is_billed,
                    invoice_num,
                    date,
                ) = entry

                try:
                    start_dt = datetime.fromisoformat(start_time)
                    start_display = start_dt.strftime("%I:%M %p")
                except Exception:
                    start_display = start_time

                try:
                    end_dt = datetime.fromisoformat(end_time)
                    end_display = end_dt.strftime("%I:%M %p")
                except Exception:
                    end_display = end_time

                duration_hours = (duration_mins / 60.0) if duration_mins else 0
                billed_status = "Yes" if is_billed else "No"

                ws.append(
                    [
                        date,
                        client,
                        project,
                        task,
                        start_display,
                        end_display,
                        round(duration_hours, 2),
                        description or "",
                        billed_status,
                        invoice_num or "",
                    ]
                )

            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except Exception:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width

            wb.save(filename)
            messagebox.showinfo("Success", f"Exported {len(entries)} time entries to:\n\n{filename}")

            if messagebox.askyesno("Open File?", "Would you like to open the exported file now?"):
                os.startfile(filename)

        except Exception as e:
            messagebox.showerror("Export Error", f"Failed to export time entries:\n\n{e}\n\n{type(e).__name__}")
            import traceback

            traceback.print_exc()

    def sync_embedded_tk_widgets(self) -> None:
        from ui.ctk.ttk_theme import embedded_tk_frame_bg

        self._tree_host.configure(bg=embedded_tk_frame_bg(), highlightthickness=0)
