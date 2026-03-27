# Version: 2026-02-03
import os
import tkinter as tk
from tkinter import ttk, messagebox, filedialog
from datetime import datetime, timedelta
from db_manager import DatabaseManager
from models import Client, Project, Task, TimeEntry, CompanyInfo
from themes import AVAILABLE_THEMES
import sqlite3
import threading
import time
import tempfile
import os
from ui_helpers import (
    center_dialog,
    center_window,
    load_theme_preference,
    restore_tree_state,
    save_theme_preference,
    save_tree_state,
)
from ui.tk.clients_tab import ClientsTabMixin
from ui.tk.company_tab import CompanyTabMixin
from ui.tk.email_tab import EmailTabMixin
from ui.tk.invoice_tab import InvoiceTabMixin
from ui.tk.projects_tab import ProjectsTabMixin
from ui.tk.tasks_tab import TasksTabMixin
from ui.tk.time_entries_tab import TimeEntriesTabMixin
from ui.tk.timer_tab import TimerTabMixin


class TimeTrackerApp(
    TimerTabMixin,
    ClientsTabMixin,
    ProjectsTabMixin,
    TasksTabMixin,
    TimeEntriesTabMixin,
    CompanyTabMixin,
    EmailTabMixin,
    InvoiceTabMixin,
):
    def __init__(self, root, db_path=None):
        """Initialize the Time Tracker application.

        Args:
            root: The tkinter root window
            db_path: Path to the database file (optional, defaults to 'time_tracker.db')
        """
        try:
            print("[DEBUG] TimeTrackerApp.__init__ starting...")
            self.root = root
            self.root.title("Freelance Timer Pro V2.0 - Professional Time & Invoice Management")
            
            # Modern window setup
            self.root.geometry("1200x800")
            self.root.minsize(600, 400)  # Much more flexible for small screens
            
            # Try to set custom icon (if exists)
            try:
                # Get absolute path to icon file relative to this script
                script_dir = os.path.dirname(os.path.abspath(__file__))
                icon_path = os.path.join(script_dir, "assets", "icon.ico")
                if os.path.exists(icon_path):
                    self.root.iconbitmap(icon_path)
            except Exception as e:
                print(f"[DEBUG] Could not set icon: {e}")  # Debug
                pass  # Use default icon if custom not available
            
            # Initialize database FIRST (needed for theme preference)
            if db_path:
                print(f"[DEBUG] Initializing DatabaseManager with path: {db_path}")
                self.db = DatabaseManager(db_path)
            else:
                print(f"[DEBUG] Initializing DatabaseManager with default path")
                self.db = DatabaseManager()
            print(f"[DEBUG] DatabaseManager initialized successfully")
            
            # Load theme (colors and fonts) AFTER database is ready
            print("[DEBUG] Loading theme...")
            saved_theme = self.load_theme_preference()
            # Get theme from registry, fallback to first available theme if not found
            self.current_theme = AVAILABLE_THEMES.get(saved_theme, list(AVAILABLE_THEMES.values())[0])
            self.colors = self.current_theme.get_colors()
            self.fonts = self.current_theme.get_fonts()
            print(f"[DEBUG] Theme loaded: {saved_theme} ({len(self.colors)} colors, {len(self.fonts)} fonts)")

            # Initialize models
            print("[DEBUG] Initializing models...")
            self.client_model = Client(self.db)
            self.project_model = Project(self.db)
            self.task_model = Task(self.db)
            self.time_entry_model = TimeEntry(self.db)
            self.company_model = CompanyInfo(self.db)
            print("[DEBUG] Models initialized")

            # Timer variables
            self.timer_running = False
            self.timer_start_time = None
            self.current_task_id = None
            
            # Daily session tracking
            self.session_date = datetime.now().date()
            self.daily_client_totals = {}  # {client_id: total_seconds}
            self.daily_project_totals = {}  # {(client_id, project_id): total_seconds}
            self.last_timer_elapsed = 0
            self.last_timer_client_id = None
            self.last_timer_project_id = None
            print("[DEBUG] Timer variables initialized")

            print("[DEBUG] Creating widgets...")
            self.create_widgets()
            print("[DEBUG] Widgets created")

            print("[DEBUG] Refreshing data...")
            self.refresh_all_data()
            print("[DEBUG] Data refreshed")
            
            # Apply modern theme and center window
            print("[DEBUG] Applying modern theme...")
            self.apply_modern_theme()
            self.center_window()
            print("[DEBUG] Theme applied")

            print("[DEBUG] TimeTrackerApp initialization complete!")

        except Exception as e:
            print(f"\n{'=' * 60}")
            print(f"ERROR IN TimeTrackerApp.__init__:")
            print(f"{'=' * 60}")
            print(f"{e}")
            import traceback
            traceback.print_exc()
            print(f"{'=' * 60}\n")
            raise
    
    def center_window(self):
        center_window(self.root)
    
    def center_dialog(self, dialog, width, height):
        center_dialog(self.root, dialog, width, height)
    
    def apply_modern_theme(self):
        """Apply theme styling to the application using theme system"""
        print("[DEBUG] Applying theme styles...")
        style = ttk.Style()
        
        # Apply current theme
        self.current_theme.apply_theme(style, self.colors, self.fonts)
        
        # Set root window background
        self.root.configure(bg=self.colors['background'])
        print("[DEBUG] Theme styles applied successfully")
    
    def load_theme_preference(self):
        return load_theme_preference(self.db.db_path, default_theme="Burnt Orange Pro V3")
    
    def save_theme_preference(self, theme_name):
        try:
            save_theme_preference(self.db.db_path, theme_name)
        except Exception as e:
            print(f"[ERROR] Could not save theme preference: {e}")
    
    def switch_theme(self, theme_name):
        """Switch to a different theme"""
        if theme_name not in AVAILABLE_THEMES:
            messagebox.showerror("Error", f"Theme '{theme_name}' not found")
            return
        
        # Load new theme
        self.current_theme = AVAILABLE_THEMES[theme_name]
        self.colors = self.current_theme.get_colors()
        self.fonts = self.current_theme.get_fonts()
        
        # Save preference
        self.save_theme_preference(theme_name)
        
        # Reapply theme
        self.apply_modern_theme()
        
        messagebox.showinfo("Theme Changed", 
                          f"Theme changed to '{theme_name}'\n\n" +
                          "Your preference has been saved.\n" +
                          "Note: Some changes may require restarting the app for full effect.")

    def save_tree_state(self, tree):
        return save_tree_state(tree)
    
    def restore_tree_state(self, tree, expanded_items, expand_all=False):
        restore_tree_state(tree, expanded_items, expand_all=expand_all)

    def create_widgets(self):
        # Create notebook for tabs
        self.notebook = ttk.Notebook(self.root)
        self.notebook.pack(fill='both', expand=True, padx=10, pady=10)

        # Create tabs
        self.create_timer_tab()
        self.create_clients_tab()
        self.create_projects_tab()
        self.create_tasks_tab()
        self.create_time_entries_tab()
        self.create_company_tab()
        self.create_invoice_tab()
        self.create_email_tab()

    # Timer methods
    def start_timer(self):
        if not self.timer_client_combo.get():
            messagebox.showerror("Error", "Please select a client first")
            return

        if not self.timer_project_combo.get():
            messagebox.showerror("Error", "Please select a project first")
            return

        if not self.timer_task_combo.get():
            messagebox.showerror("Error", "Please select a task first")
            return

        # Get task ID from the selected display text
        task_text = self.timer_task_combo.get()
        client_name = self.timer_client_combo.get()
        project_name = self.timer_project_combo.get()

        # Check if it's a global task
        if task_text.startswith('[GLOBAL] '):
            # Extract task name from "[GLOBAL] TaskName"
            task_name = task_text.replace('[GLOBAL] ', '')

            # Find the global task
            global_tasks = self.task_model.get_global_tasks()
            self.current_task_id = None

            for task in global_tasks:
                if task[2] == task_name:
                    self.current_task_id = task[0]
                    break
        else:
            # Regular project-specific task
            # Extract task name from "Client - Project - Task" format
            parts = task_text.split(' - ')
            if len(parts) >= 3:
                task_name = ' - '.join(parts[2:])
            else:
                messagebox.showerror("Error", "Invalid task format")
                return

            # Find the task by matching client, project, and task names
            tasks = self.task_model.get_all()
            self.current_task_id = None

            for task in tasks:
                if task[9] == client_name and task[8] == project_name and task[2] == task_name:
                    self.current_task_id = task[0]
                    break

        if not self.current_task_id:
            messagebox.showerror("Error", f"Could not find task: {task_name}")
            return

        try:
            # Start the timer
            self.timer_running = True
            self.timer_start_time = datetime.now()

            # Get current project ID for global tasks
            project_id = self.get_current_timer_project_id()
            self.time_entry_model.start_timer(self.current_task_id, project_id_override=project_id)

            # Update UI
            self.start_button.config(state='disabled')
            self.stop_button.config(state='normal')
            self.timer_label.config(text="00:00:00", foreground='green')  # Reset and show green

            # Start timer thread
            self.update_timer_display()
        except Exception as e:
            messagebox.showerror("Error", f"Failed to start timer: {str(e)}")
            self.timer_running = False

    def stop_timer(self):
        if self.timer_running:
            # Calculate elapsed time BEFORE stopping
            elapsed_seconds = (datetime.now() - self.timer_start_time).total_seconds()
            
            # Get current client and project IDs for tracking
            client_id = self.get_current_timer_client_id()
            project_id = self.get_current_timer_project_id()
            
            # Stop the timer in database
            self.timer_running = False
            self.time_entry_model.stop_timer()

            # Update daily totals (both client and project)
            if client_id:
                if client_id not in self.daily_client_totals:
                    self.daily_client_totals[client_id] = 0
                self.daily_client_totals[client_id] += elapsed_seconds
                self.last_timer_client_id = client_id
                
                # Track project totals
                if project_id:
                    key = (client_id, project_id)
                    if key not in self.daily_project_totals:
                        self.daily_project_totals[key] = 0
                    self.daily_project_totals[key] += elapsed_seconds
                    self.last_timer_project_id = project_id
            
            # Store elapsed time for display
            self.last_timer_elapsed = elapsed_seconds
            
            # Update UI with FINAL time (don't reset to zero!)
            self.start_button.config(state='normal')
            self.stop_button.config(state='disabled')
            self.update_timer_display_final()
            
            # Update daily totals display
            self.update_daily_totals_display()

            self.refresh_time_entries()
            
            # Show success with elapsed time
            hours = int(elapsed_seconds // 3600)
            minutes = int((elapsed_seconds % 3600) // 60)
            seconds = int(elapsed_seconds % 60)
            messagebox.showinfo("Success", 
                              f"Timer stopped and time entry saved\n\n" +
                              f"Time recorded: {hours:02d}:{minutes:02d}:{seconds:02d}")

    def update_timer_display(self):
        if self.timer_running and self.timer_start_time:
            elapsed = datetime.now() - self.timer_start_time
            hours, remainder = divmod(int(elapsed.total_seconds()), 3600)
            minutes, seconds = divmod(remainder, 60)
            self.timer_label.config(text=f"{hours:02d}:{minutes:02d}:{seconds:02d}", foreground='green')
            self.root.after(1000, self.update_timer_display)
    
    def update_timer_display_final(self):
        """Display final elapsed time after stopping (don't reset to zero)"""
        if self.last_timer_elapsed > 0:
            hours = int(self.last_timer_elapsed // 3600)
            minutes = int((self.last_timer_elapsed % 3600) // 60)
            seconds = int(self.last_timer_elapsed % 60)
            
            self.timer_label.config(
                text=f"Last: {hours:02d}:{minutes:02d}:{seconds:02d}",
                foreground='gray'
            )
        else:
            self.timer_label.config(text="00:00:00", foreground='black')

    def on_timer_client_select(self, event):
        client_name = self.timer_client_combo.get()
        if client_name:
            # Get client ID
            clients = self.client_model.get_all()
            client_id = None
            for client in clients:
                if client[1] == client_name:
                    client_id = client[0]
                    break

            if client_id:
                # Load projects for this client
                projects = self.project_model.get_by_client(client_id)
                self.timer_project_combo['values'] = [p[2] for p in projects]
                self.timer_project_combo.set('')
                self.timer_task_combo.set('')

    def on_timer_project_select(self, event):
        project_name = self.timer_project_combo.get()
        client_name = self.timer_client_combo.get()

        if project_name and client_name:
            # Get project ID by matching client and project names
            projects = self.project_model.get_all()
            project_id = None
            for project in projects:
                proj_client_name = project[9] if len(project) > 9 else None
                if proj_client_name == client_name and project[2] == project_name:
                    project_id = project[0]
                    break

            if project_id:
                # Load tasks for this project AND global tasks
                project_tasks = self.task_model.get_by_project(project_id)
                global_tasks = self.task_model.get_global_tasks()

                # Combine both lists
                all_tasks = list(global_tasks) + list(project_tasks)

                # Format task displays
                task_displays = []
                for t in all_tasks:
                    if t[1] is None:  # project_id is None means it's global
                        task_displays.append(f"[GLOBAL] {t[2]}")
                    else:
                        task_displays.append(f"{client_name} - {project_name} - {t[2]}")

                self.timer_task_combo['values'] = task_displays
                self.timer_task_combo.set('')
            else:
                self.timer_task_combo['values'] = []
                self.timer_task_combo.set('')


    def get_current_timer_client_id(self):
        """Get the client ID for the currently selected timer task"""
        try:
            client_name = self.timer_client_combo.get()
            if not client_name:
                return None
            
            clients = self.client_model.get_all()
            for client in clients:
                if client[1] == client_name:
                    return client[0]
            return None
        except:
            return None
    
    def get_current_timer_project_id(self):
        """Get the project ID for the currently selected timer project"""
        try:
            client_name = self.timer_client_combo.get()
            project_name = self.timer_project_combo.get()
            
            if not client_name or not project_name:
                return None
            
            # Get all projects and find matching one
            projects = self.project_model.get_all()
            for project in projects:
                # project structure: [0:id, ..., 9:client_name]
                proj_client_name = project[9] if len(project) > 9 else None
                if proj_client_name == client_name and project[2] == project_name:
                    return project[0]
            return None
        except:
            return None
    
    def update_daily_totals_display(self):
        """Update the daily totals display showing time accumulated per client with project breakdowns"""
        # Check if we need to reset for new day
        current_date = datetime.now().date()
        if current_date != self.session_date:
            self.daily_client_totals = {}
            self.daily_project_totals = {}
            self.session_date = current_date
        
        # Enable BOTH text widgets for editing
        self.daily_totals_text.config(state='normal')
        self.daily_totals_text.delete('1.0', 'end')
        self.manual_daily_totals_text.config(state='normal')
        self.manual_daily_totals_text.delete('1.0', 'end')
        
        if not self.daily_client_totals:
            text = (f"📊 Daily Time Tracker - {self.session_date.strftime('%B %d, %Y')}\n\n" +
                "No time tracked yet today.\n\n" +
                "Start a timer to begin tracking!")
            self.daily_totals_text.insert('1.0', text)
            self.manual_daily_totals_text.insert('1.0', text)
        else:
            # Build header
            text = f"📊 Daily Time Tracker - {self.session_date.strftime('%B %d, %Y')}\n"
            text += "=" * 50 + "\n\n"
            
            # Calculate grand total
            grand_total_seconds = sum(self.daily_client_totals.values())
            grand_total_hours = grand_total_seconds / 3600
            
            # Get client and project names
            clients = self.client_model.get_all()
            client_dict = {c[0]: c[1] for c in clients}
            
            projects = self.project_model.get_all()
            project_dict = {p[0]: p[2] for p in projects}  # {project_id: project_name}
            
            # Group projects by client for display
            client_projects = {}  # {client_id: [(project_id, seconds), ...]}
            for (client_id, project_id), seconds in self.daily_project_totals.items():
                if client_id not in client_projects:
                    client_projects[client_id] = []
                client_projects[client_id].append((project_id, seconds))
            
            # Display each client with project breakdowns
            for client_id, total_seconds in sorted(self.daily_client_totals.items(), 
                                                   key=lambda x: x[1], reverse=True):
                client_name = client_dict.get(client_id, f"Client #{client_id}")
                hours = total_seconds / 3600
                
                # Format client total as HH:MM:SS and decimal hours
                h = int(total_seconds // 3600)
                m = int((total_seconds % 3600) // 60)
                s = int(total_seconds % 60)
                
                text += f"📌 {client_name}:  {h:02d}:{m:02d}:{s:02d}  ({hours:.2f} hrs)\n"
                
                # Show project breakdown if available
                if client_id in client_projects:
                    # Sort projects by time (most first)
                    sorted_projects = sorted(client_projects[client_id], 
                                           key=lambda x: x[1], reverse=True)
                    
                    for project_id, proj_seconds in sorted_projects:
                        project_name = project_dict.get(project_id, f"Project #{project_id}")
                        proj_hours = proj_seconds / 3600
                        
                        # Format project time
                        ph = int(proj_seconds // 3600)
                        pm = int((proj_seconds % 3600) // 60)
                        ps = int(proj_seconds % 60)
                        
                        text += f"    └─ {project_name}:  {ph:02d}:{pm:02d}:{ps:02d}  ({proj_hours:.2f} hrs)\n"
                
                text += "\n"  # Blank line between clients
            
            # Add grand total
            text += "=" * 50 + "\n"
            h = int(grand_total_seconds // 3600)
            m = int((grand_total_seconds % 3600) // 60)
            s = int(grand_total_seconds % 60)
            text += f"  TOTAL TODAY:  {h:02d}:{m:02d}:{s:02d}  ({grand_total_hours:.2f} hrs)"
            
            self.daily_totals_text.insert('1.0', text)
            self.manual_daily_totals_text.insert('1.0', text)
        
        # Disable editing on BOTH widgets
        self.daily_totals_text.config(state='disabled')
        self.manual_daily_totals_text.config(state='disabled')
    
    def reset_daily_totals(self):
        """Manually reset daily totals"""
        if not self.daily_client_totals:
            messagebox.showinfo("Nothing to Reset", "No daily totals to reset.")
            return
        
        response = messagebox.askyesno(
            "Reset Daily Totals",
            "Are you sure you want to reset today's accumulated time?\n\n" +
            "This will only reset the daily display tracker.\n" +
            "Saved time entries will not be affected."
        )
        
        if response:
            self.daily_client_totals = {}
            self.daily_project_totals = {}
            self.last_timer_elapsed = 0
            self.update_daily_totals_display()
            self.timer_label.config(text="00:00:00", foreground='black')
            messagebox.showinfo("Reset Complete", "Daily totals have been reset.")
    
    def on_task_client_select(self, event):
        """When client is selected in tasks tab, populate projects"""
        client_name = self.task_client_combo.get()
        if client_name:
            # Get client ID
            clients = self.client_model.get_all()
            client_id = None
            for client in clients:
                if client[1] == client_name:
                    client_id = client[0]
                    break

            if client_id:
                # Load projects for this client
                projects = self.project_model.get_by_client(client_id)
                self.task_project_combo['values'] = [p[2] for p in projects]
                self.task_project_combo.set('')

    def toggle_manual_entry_mode(self):
        """Toggle between time range and decimal hour entry modes"""
        mode = self.manual_entry_mode.get()
        
        if mode == "time_range":
            # Show start/end time fields
            self.manual_start_entry.grid()
            self.manual_end_entry.grid()
            
            # Hide decimal fields
            self.manual_decimal_label.grid_remove()
            self.manual_decimal_entry.grid_remove()
            self.manual_decimal_help.grid_remove()
        else:  # decimal
            # Hide start/end time fields
            self.manual_start_entry.grid_remove()
            self.manual_end_entry.grid_remove()
            
            # Show decimal fields
            self.manual_decimal_label.grid()
            self.manual_decimal_entry.grid()
            self.manual_decimal_help.grid()

    def add_manual_entry(self):
        task_text = self.manual_task_combo.get()
        date_str = self.manual_date_entry.get().strip()
        description = self.manual_desc_text.get("1.0", tk.END).strip()
        mode = self.manual_entry_mode.get()

        if not task_text:
            messagebox.showerror("Error", "Please select a task")
            return

        try:
            # Parse date MM/DD/YY
            date_obj = datetime.strptime(date_str, "%m/%d/%y")

            if mode == "time_range":
                # Original time range mode
                start_str = self.manual_start_entry.get().strip()
                end_str = self.manual_end_entry.get().strip()

                # Parse times HH:MM AM/PM
                start_time_obj = datetime.strptime(f"{date_str} {start_str}", "%m/%d/%y %I:%M %p")
                end_time_obj = datetime.strptime(f"{date_str} {end_str}", "%m/%d/%y %I:%M %p")

                if end_time_obj <= start_time_obj:
                    messagebox.showerror("Error", "End time must be after start time")
                    return

            else:  # decimal mode
                # Parse decimal hours
                decimal_str = self.manual_decimal_entry.get().strip()

                if not decimal_str:
                    messagebox.showerror("Error", "Please enter hours in decimal format")
                    return

                try:
                    decimal_hours = float(decimal_str)

                    if decimal_hours <= 0:
                        messagebox.showerror("Error", "Hours must be greater than 0")
                        return

                    if decimal_hours > 24:
                        messagebox.showerror("Error", "Hours cannot exceed 24 in a single entry")
                        return

                except ValueError:
                    messagebox.showerror("Error",
                                         f"Invalid decimal format: '{decimal_str}'\n\n" +
                                         "Examples: 1.5, 0.75, 2.25")
                    return

                # Calculate start and end times from decimal hours
                # Use 9 AM as default start time for decimal entries
                start_time_obj = datetime.strptime(f"{date_str} 09:00 AM", "%m/%d/%y %I:%M %p")

                # Calculate end time by adding decimal hours
                duration = timedelta(hours=decimal_hours)
                end_time_obj = start_time_obj + duration

        except ValueError as e:
            messagebox.showerror("Error",
                                 f"Invalid date/time format.\nUse MM/DD/YY for date and HH:MM AM/PM for time.\nError: {str(e)}")
            return

        # Get task ID (handle both global and regular tasks)
        task_id = None

        # Check if it's a global task
        if task_text.startswith('[GLOBAL] '):
            task_name = task_text.replace('[GLOBAL] ', '')
            global_tasks = self.task_model.get_global_tasks()
            for task in global_tasks:
                if task[2] == task_name:
                    task_id = task[0]
                    break
        else:
            # Regular project task
            tasks = self.task_model.get_all()
            for task in tasks:
                client_name = task[9]
                project_name = task[8]
                task_display = f"{client_name} - {project_name} - {task[2]}"
                if task_display == task_text:
                    task_id = task[0]
                    break

        if not task_id:
            messagebox.showerror("Error", "Invalid task selected")
            return

        # Get project_id for global tasks
        project_id_override = None
        if task_text.startswith('[GLOBAL] '):
            project_id_override = self.get_manual_entry_project_id()
            
        self.time_entry_model.add_manual_entry(task_id, start_time_obj, end_time_obj, description, project_id_override=project_id_override)
        self.refresh_time_entries()

        # Calculate duration for display and daily totals
        duration = end_time_obj - start_time_obj
        hours = duration.total_seconds() / 3600
        duration_seconds = duration.total_seconds()

        # Update daily totals if entry is for today
        entry_date = start_time_obj.date()
        today = datetime.now().date()

        if entry_date == today:
            # Get client and project IDs from the manual entry form (works for both global and regular tasks)
            client_id = None
            project_id = None
            
            # Get client ID from manual entry form
            client_name = self.manual_client_combo.get()
            if client_name:
                clients = self.client_model.get_all()
                for client in clients:
                    if client[1] == client_name:
                        client_id = client[0]
                        break
            
            # Get project ID from manual entry form
            project_name = self.manual_project_combo.get()
            if project_name and client_id:
                projects = self.project_model.get_by_client(client_id)
                for project in projects:
                    if project[2] == project_name:
                        project_id = project[0]
                        break
            
            # Update daily totals if we found both client and project
            if client_id and project_id:
                # Update client totals
                if client_id not in self.daily_client_totals:
                    self.daily_client_totals[client_id] = 0
                self.daily_client_totals[client_id] += duration_seconds

                # Update project totals
                key = (client_id, project_id)
                if key not in self.daily_project_totals:
                    self.daily_project_totals[key] = 0
                self.daily_project_totals[key] += duration_seconds

            # Refresh the daily totals display
            self.update_daily_totals_display()

        # Show success message with details
        if mode == "decimal":
            messagebox.showinfo("Success",
                                f"Time entry added successfully\n\n" +
                                f"Duration: {decimal_hours} hours\n" +
                                f"Date: {date_obj.strftime('%m/%d/%y')}")
        else:
            messagebox.showinfo("Success",
                                f"Time entry added successfully\n\n" +
                                f"Duration: {hours:.2f} hours\n" +
                                f"From: {start_time_obj.strftime('%I:%M %p')} to {end_time_obj.strftime('%I:%M %p')}")

        self.clear_manual_entry_form()

    def on_manual_client_select(self, event):
        """When client is selected in manual entry, populate projects"""
        client_name = self.manual_client_combo.get()
        if client_name:
            # Get client ID
            clients = self.client_model.get_all()
            client_id = None
            for client in clients:
                if client[1] == client_name:
                    client_id = client[0]
                    break

            if client_id:
                projects = self.project_model.get_by_client(client_id)
                self.manual_project_combo['values'] = [p[2] for p in projects]
                self.manual_project_combo.set('')
                self.manual_task_combo.set('')

    def on_manual_project_select(self, event):
        """When project is selected in manual entry, populate tasks"""
        project_name = self.manual_project_combo.get()
        client_name = self.manual_client_combo.get()

        if project_name and client_name:
            projects = self.project_model.get_all()
            project_id = None
            for project in projects:
                proj_client_name = project[9] if len(project) > 9 else None
                if proj_client_name == client_name and project[2] == project_name:
                    project_id = project[0]
                    break

            if project_id:
                project_tasks = self.task_model.get_by_project(project_id)
                global_tasks = self.task_model.get_global_tasks()
                all_tasks = list(global_tasks) + list(project_tasks)
                
                task_displays = []
                for t in all_tasks:
                    if t[1] is None:
                        task_displays.append(f"[GLOBAL] {t[2]}")
                    else:
                        task_displays.append(f"{client_name} - {project_name} - {t[2]}")

                self.manual_task_combo['values'] = task_displays
                self.manual_task_combo.set('')

    def get_manual_entry_project_id(self):
        """Get the project ID from manual entry form"""
        try:
            client_name = self.manual_client_combo.get()
            project_name = self.manual_project_combo.get()
            
            if not client_name or not project_name:
                return None
            
            projects = self.project_model.get_all()
            for project in projects:
                proj_client_name = project[9] if len(project) > 9 else None
                if proj_client_name == client_name and project[2] == project_name:
                    return project[0]
            return None
        except:
            return None

    def clear_manual_entry_form(self):
        self.manual_date_entry.delete(0, tk.END)
        self.manual_date_entry.insert(0, datetime.now().strftime("%m/%d/%y"))
        self.manual_start_entry.delete(0, tk.END)
        self.manual_start_entry.insert(0, "09:00 AM")
        self.manual_client_combo.set('')
        self.manual_project_combo.set('')
        self.manual_end_entry.delete(0, tk.END)
        self.manual_end_entry.insert(0, "05:00 PM")
        self.manual_decimal_entry.delete(0, tk.END)
        self.manual_task_combo.set('')
        self.manual_desc_text.delete("1.0", tk.END)

    # Client methods
    def add_client(self):
        name = self.client_name_entry.get().strip()
        company = self.client_company_entry.get().strip()
        email = self.client_email_entry.get().strip()
        phone = self.client_phone_entry.get().strip()
        address = self.client_address_text.get("1.0", tk.END).strip()

        if not name:
            messagebox.showerror("Error", "Client name is required")
            return

        # CHECK FOR DUPLICATE CLIENT NAME
        existing_clients = self.client_model.get_all()
        for client in existing_clients:
            if client[1].lower() == name.lower():  # Case-insensitive comparison
                messagebox.showerror("Error", f"Client '{name}' already exists")
                return

        self.client_model.create(name, company, email, phone, address)
        self.clear_client_form()
        self.refresh_clients()
        self.refresh_combos()
        messagebox.showinfo("Success", "Client added successfully")

    def update_client(self):
        selection = self.client_tree.selection()
        if not selection:
            messagebox.showerror("Error", "Please select a client to update")
            return

        client_id = self.client_tree.item(selection[0])['values'][0]
        name = self.client_name_entry.get().strip()
        company = self.client_company_entry.get().strip()
        email = self.client_email_entry.get().strip()
        phone = self.client_phone_entry.get().strip()
        address = self.client_address_text.get("1.0", tk.END).strip()

        if not name:
            messagebox.showerror("Error", "Client name is required")
            return

        self.client_model.update(client_id, name, company, email, phone, address)
        self.clear_client_form()
        self.refresh_clients()
        self.refresh_combos()
        messagebox.showinfo("Success", "Client updated successfully")

    def delete_client(self):
        selection = self.client_tree.selection()
        if not selection:
            messagebox.showerror("Error", "Please select a client to delete")
            return

        if messagebox.askyesno("Confirm",
                               "Delete this client? This will also delete all associated projects, tasks, and time entries."):
            client_id = self.client_tree.item(selection[0])['values'][0]
            self.client_model.delete(client_id)
            self.refresh_clients()
            self.refresh_combos()
            messagebox.showinfo("Success", "Client deleted successfully")

    def clear_client_form(self):
        self.client_name_entry.delete(0, tk.END)
        self.client_company_entry.delete(0, tk.END)
        self.client_email_entry.delete(0, tk.END)
        self.client_phone_entry.delete(0, tk.END)
        self.client_address_text.delete("1.0", tk.END)

    def on_client_select(self, event):
        selection = self.client_tree.selection()
        if selection:
            client_id = self.client_tree.item(selection[0])['values'][0]
            client = self.client_model.get_by_id(client_id)
            if client:
                self.client_name_entry.delete(0, tk.END)
                self.client_name_entry.insert(0, client[1])

                self.client_company_entry.delete(0, tk.END)
                self.client_company_entry.insert(0, client[2] or "")

                self.client_email_entry.delete(0, tk.END)
                self.client_email_entry.insert(0, client[3] or "")

                self.client_phone_entry.delete(0, tk.END)
                self.client_phone_entry.insert(0, client[4] or "")

                self.client_address_text.delete("1.0", tk.END)
                self.client_address_text.insert("1.0", client[5] or "")

    # Project methods
    def add_project(self):
        client_text = self.project_client_combo.get()
        name = self.project_name_entry.get().strip()
        description = self.project_desc_text.get("1.0", tk.END).strip()

        if not client_text or not name:
            messagebox.showerror("Error", "Client and project name are required")
            return

        # Get client ID by name
        clients = self.client_model.get_all()
        client_id = None
        for client in clients:
            if client[1] == client_text:
                client_id = client[0]
                break

        if not client_id:
            messagebox.showerror("Error", "Invalid client selected")
            return

        # CHECK FOR DUPLICATE PROJECT NAME UNDER SAME CLIENT
        existing_projects = self.project_model.get_by_client(client_id)
        for proj in existing_projects:
            if proj[2].lower() == name.lower():  # Case-insensitive comparison
                messagebox.showerror("Error", f"Project '{name}' already exists for this client")
                return

        rate = float(self.project_rate_entry.get() or 0)
        is_lump_sum = self.project_billing_var.get() == "lump_sum"

        if is_lump_sum:
            self.project_model.create(client_id, name, description, 0, True, rate)
        else:
            self.project_model.create(client_id, name, description, rate, False, 0)

        self.clear_project_form()
        self.refresh_projects()
        self.refresh_combos()
        messagebox.showinfo("Success", "Project added successfully")

    def update_project(self):
        selection = self.project_tree.selection()
        if not selection:
            messagebox.showerror("Error", "Please select a project to update")
            return

        project_id = self.project_tree.item(selection[0])['values'][0]
        client_text = self.project_client_combo.get()
        name = self.project_name_entry.get().strip()
        description = self.project_desc_text.get("1.0", tk.END).strip()

        if not client_text or not name:
            messagebox.showerror("Error", "Client and project name are required")
            return

        # Get client ID by name
        clients = self.client_model.get_all()
        client_id = None
        for client in clients:
            if client[1] == client_text:
                client_id = client[0]
                break

        if not client_id:
            messagebox.showerror("Error", "Invalid client selected")
            return

        rate = float(self.project_rate_entry.get() or 0)
        is_lump_sum = self.project_billing_var.get() == "lump_sum"

        if is_lump_sum:
            self.project_model.update(project_id, client_id, name, description, 0, True, rate)
        else:
            self.project_model.update(project_id, client_id, name, description, rate, False, 0)

        self.clear_project_form()
        self.refresh_projects()
        self.refresh_combos()
        messagebox.showinfo("Success", "Project updated successfully")

    def delete_project(self):
        selection = self.project_tree.selection()
        if not selection:
            messagebox.showerror("Error", "Please select a project to delete")
            return

        if messagebox.askyesno("Confirm",
                               "Delete this project? This will also delete all associated tasks and time entries."):
            project_id = self.project_tree.item(selection[0])['values'][0]
            self.project_model.delete(project_id)
            self.refresh_projects()
            self.refresh_combos()
            messagebox.showinfo("Success", "Project deleted successfully")

    def clear_project_form(self):
        self.project_client_combo.set("")
        self.project_name_entry.delete(0, tk.END)
        self.project_desc_text.delete("1.0", tk.END)
        self.project_rate_entry.delete(0, tk.END)
        self.project_billing_var.set("hourly")

    def on_project_select(self, event):
        selection = self.project_tree.selection()
        if selection:
            project_id = self.project_tree.item(selection[0])['values'][0]
            project = self.project_model.get_by_id(project_id)
            if project:
                # Set client
                client = self.client_model.get_by_id(project[1])
                if client:
                    self.project_client_combo.set(client[1])

                # Set project details
                self.project_name_entry.delete(0, tk.END)
                self.project_name_entry.insert(0, project[2])

                self.project_desc_text.delete("1.0", tk.END)
                self.project_desc_text.insert("1.0", project[3] or "")

                if project[5]:  # is_lump_sum
                    self.project_billing_var.set("lump_sum")
                    self.project_rate_entry.delete(0, tk.END)
                    self.project_rate_entry.insert(0, str(project[6]))
                else:
                    self.project_billing_var.set("hourly")
                    self.project_rate_entry.delete(0, tk.END)
                    self.project_rate_entry.insert(0, str(project[4]))

    def toggle_project_billing(self):
        pass

    # Task methods
    def add_task(self):
        name = self.task_name_entry.get().strip()
        description = self.task_desc_text.get("1.0", tk.END).strip()
        is_global = self.task_global_var.get()

        if not name:
            messagebox.showerror("Error", "Please enter a task name")
            return

        if is_global:
            # Create global task (project_id will be None)
            rate = float(self.task_rate_entry.get() or 0)
            is_lump_sum_flag = self.task_billing_var.get() == "lump_sum"

            if is_lump_sum_flag:
                self.task_model.create(name, description, 0, True, rate, None, True)
            else:
                self.task_model.create(name, description, rate, False, 0, None, True)

            messagebox.showinfo("Success", f"Global task '{name}' created!")
        else:
            # Create project-specific task
            client_text = self.task_client_combo.get()
            project_text = self.task_project_combo.get()

            if not client_text or not project_text:
                messagebox.showerror("Error", "Please select a client and project for non-global tasks")
                return

            # Get project ID by matching client and project names
            projects = self.project_model.get_all()
            project_id = None
            for project in projects:
                client_name = project[9] if len(project) > 9 else None
                if client_name == client_text and project[2] == project_text:
                    project_id = project[0]
                    break

            if not project_id:
                messagebox.showerror("Error", "Invalid project selected")
                return

            # CHECK FOR DUPLICATE TASK NAME UNDER SAME PROJECT
            existing_tasks = self.task_model.get_by_project(project_id)
            for task in existing_tasks:
                if task[2].lower() == name.lower():
                    messagebox.showerror("Error", f"Task '{name}' already exists for this project")
                    return

            rate = float(self.task_rate_entry.get() or 0)
            is_lump_sum_flag = self.task_billing_var.get() == "lump_sum"

            if is_lump_sum_flag:
                self.task_model.create(name, description, 0, True, rate, project_id, False)
            else:
                self.task_model.create(name, description, rate, False, 0, project_id, False)

            messagebox.showinfo("Success", f"Task '{name}' created!")

        self.clear_task_form()
        self.refresh_tasks()
        self.refresh_combos()


    def update_task(self):
        selection = self.task_tree.selection()
        if not selection:
            messagebox.showerror("Error", "Please select a task to update")
            return

        # Extract task ID from tags (same as on_task_select)
        item = selection[0]
        item_tags = self.task_tree.item(item)['tags']
        item_values = self.task_tree.item(item)['values']
        
        # Only allow updating actual tasks (not client/project headers)
        if 'task' not in item_tags:
            messagebox.showerror("Error", "Please select an actual task (not a client/project group)")
            return
        
        # Extract task ID from tags
        task_id = None
        for tag in item_tags:
            if tag.startswith('task_id_'):
                task_id = int(tag.replace('task_id_', ''))
                break
        
        # Fallback: try values column
        if not task_id and len(item_values) > 1:
            try:
                task_id = int(item_values[1])
            except:
                pass
        
        if not task_id:
            messagebox.showerror("Error", "Could not determine task ID")
            return

        name = self.task_name_entry.get().strip()
        description = self.task_desc_text.get("1.0", tk.END).strip()

        if not name:
            messagebox.showerror("Error", "Task name is required")
            return

        rate = float(self.task_rate_entry.get() or 0)
        is_lump_sum = self.task_billing_var.get() == "lump_sum"

        if is_lump_sum:
            self.task_model.update(task_id, name, description, 0, True, rate)
        else:
            self.task_model.update(task_id, name, description, rate, False, 0)

        self.clear_task_form()
        self.refresh_tasks()
        self.refresh_combos()
        messagebox.showinfo("Success", "Task updated successfully")

    def delete_task(self):
        selection = self.task_tree.selection()
        if not selection:
            messagebox.showerror("Error", "Please select a task to delete")
            return
        
        # Get task ID from tags (not values which might be header)
        item_tags = self.task_tree.item(selection[0])['tags']
        if 'task' not in item_tags:
            messagebox.showerror("Error", "Please select an actual task (not a client/project header)")
            return
        
        # Extract task ID from tags
        task_id = None
        for tag in item_tags:
            if tag.startswith('task_id_'):
                task_id = int(tag.replace('task_id_', ''))
                break
        
        if not task_id:
            messagebox.showerror("Error", "Could not determine task ID")
            return

        if messagebox.askyesno("Confirm", "Delete this task? This will also delete all associated time entries."):
            self.task_model.delete(task_id)
            self.refresh_tasks()
            self.refresh_combos()
            messagebox.showinfo("Success", "Task deleted successfully")

    def clear_task_form(self):
        self.task_client_combo.set("")
        self.task_project_combo.set("")
        self.task_name_entry.delete(0, tk.END)
        self.task_desc_text.delete("1.0", tk.END)
        self.task_rate_entry.delete(0, tk.END)
        self.task_billing_var.set("hourly")

    def on_task_select(self, event):
        selection = self.task_tree.selection()
        if not selection:
            return
        
        # Get item tags and values
        item = selection[0]
        item_tags = self.task_tree.item(item)['tags']
        item_values = self.task_tree.item(item)['values']
        
        # Only process if this is an actual task (not client/project header)
        if 'task' not in item_tags:
            return
        
        # Extract task ID from tags (stored as 'task_id_123')
        task_id = None
        for tag in item_tags:
            if tag.startswith('task_id_'):
                task_id = int(tag.replace('task_id_', ''))
                break
        
        # Fallback: try to get ID from values column
        if not task_id and len(item_values) > 1:
            try:
                task_id = int(item_values[1])  # ID is in column 1
            except:
                pass
        
        if not task_id:
            return
        
        task = self.task_model.get_by_id(task_id)
        if not task:
            return
        
        # Check if it's a global task
        is_global = task[1] is None  # project_id is None for global tasks
        
        if is_global:
            # Set global checkbox
            self.task_global_var.set(True)
            self.toggle_task_project_field()  # This will disable client/project combos
        else:
            # Clear global checkbox
            self.task_global_var.set(False)
            self.toggle_task_project_field()  # This will enable client/project combos
            
            # Set project and client
            project = self.project_model.get_by_id(task[1])
            if project:
                client = self.client_model.get_by_id(project[1])
                if client:
                    self.task_client_combo.set(client[1])
                    # Trigger the client selection to populate projects
                    projects = self.project_model.get_by_client(client[0])
                    self.task_project_combo['values'] = [p[2] for p in projects]
                    self.task_project_combo.set(project[2])
        
        # Set task details
        self.task_name_entry.delete(0, tk.END)
        self.task_name_entry.insert(0, task[2])
        
        self.task_desc_text.delete("1.0", tk.END)
        self.task_desc_text.insert("1.0", task[3] or "")
        
        if task[5]:  # is_lump_sum
            self.task_billing_var.set("lump_sum")
            self.task_rate_entry.delete(0, tk.END)
            self.task_rate_entry.insert(0, str(task[6]))
        else:
            self.task_billing_var.set("hourly")
            self.task_rate_entry.delete(0, tk.END)
            self.task_rate_entry.insert(0, str(task[4]))

    def toggle_task_billing(self):
        pass

    
    def toggle_task_project_field(self):
        """Enable or disable project selection based on global checkbox"""
        is_global = self.task_global_var.get()
        if is_global:
            self.task_client_combo.set('')
            self.task_client_combo.config(state='disabled')
            self.task_project_combo.set('')
            self.task_project_combo.config(state='disabled')
        else:
            self.task_client_combo.config(state='readonly')
            self.task_project_combo.config(state='readonly')

    # Time Entry methods
    def edit_time_entry(self):
        selection = self.entries_tree.selection()
        if not selection:
            messagebox.showerror("Error", "Please select a time entry to edit")
            return
        
        # Get the selected item
        selected_item = selection[0]
        item_tags = self.entries_tree.item(selected_item)['tags']
        
        # Only allow editing of actual entries, not group nodes
        if 'entry' not in item_tags:
            messagebox.showerror("Error", "Please select an individual time entry (not a group)")
            return
        
        # Extract entry ID from tags
        entry_id = None
        for tag in item_tags:
            if tag.startswith('entry_id_'):
                entry_id = int(tag.replace('entry_id_', ''))
                break
        
        if not entry_id:
            messagebox.showerror("Error", "Could not find entry ID")
            return

        self.open_edit_time_entry_dialog(entry_id)

    def open_edit_time_entry_dialog_and_refresh(self, entry_id):
        """Open edit dialog and refresh invoice list after closing"""
        edit_window = tk.Toplevel(self.root)
        edit_window.title("Edit Time Entry")
        self.center_dialog(edit_window, 500, 400)
        
        # Make dialog modal and bring to front
        edit_window.transient(self.root)
        edit_window.grab_set()
        
        # Center on parent window
        edit_window.update_idletasks()
        x = self.root.winfo_x() + (self.root.winfo_width() // 2) - (edit_window.winfo_width() // 2)
        y = self.root.winfo_y() + (self.root.winfo_height() // 2) - (edit_window.winfo_height() // 2)
        edit_window.geometry(f"+{x}+{y}")
        
        edit_window.lift()
        edit_window.focus_force()
        
        # Refresh when dialog closes
        def on_close():
            edit_window.destroy()
            # Reload invoice entries to show changes
            if hasattr(self, 'invoice_entries_tree'):
                self.load_invoiceable_entries()
        
        edit_window.protocol("WM_DELETE_WINDOW", on_close)
        
        # Continue with rest of dialog creation
        self._build_edit_dialog_content(edit_window, entry_id, on_close)
    
    def open_edit_time_entry_dialog(self, entry_id):
        """Original edit dialog (used by Time Entries tab)"""
        edit_window = tk.Toplevel(self.root)
        edit_window.title("Edit Time Entry")
        self.center_dialog(edit_window, 500, 400)
        
        # Make dialog appear in front
        edit_window.transient(self.root)
        
        # Center on parent window
        edit_window.update_idletasks()
        x = self.root.winfo_x() + (self.root.winfo_width() // 2) - (edit_window.winfo_width() // 2)
        y = self.root.winfo_y() + (self.root.winfo_height() // 2) - (edit_window.winfo_height() // 2)
        edit_window.geometry(f"+{x}+{y}")
        
        edit_window.lift()
        edit_window.focus_force()
        
        # Build dialog content
        self._build_edit_dialog_content(edit_window, entry_id, lambda: edit_window.destroy())
    
    def _build_edit_dialog_content(self, edit_window, entry_id, close_callback):

        # Get the entry details - simplified query using denormalized columns
        with self.db.get_connection() as conn:
            cursor = conn.cursor()
            cursor.execute('SELECT * FROM time_entries WHERE id = ?', (entry_id,))
            entry = cursor.fetchone()

        if not entry:
            messagebox.showerror("Error", f"Time entry #{entry_id} not found in database")
            close_callback()
            return

        # Create form
        form_frame = ttk.Frame(edit_window)
        form_frame.pack(fill='both', expand=True, padx=15, pady=15)

        # Get names from denormalized columns in time_entries
        # Schema: id, task_id, task_name, project_id, project_name, client_id, client_name, date, start_time...
        task_name = entry[2] if entry[2] else "Unknown"
        project_name = entry[4] if entry[4] else "Unknown"
        client_name = entry[6] if entry[6] else "Unknown"

        ttk.Label(form_frame, text=f"Task: {client_name} - {project_name} - {task_name}", 
                 font=('Arial', 9, 'bold')).grid(row=0, column=0, columnspan=2, sticky='w', pady=(0, 10))

        # Get original start and end times (columns 8 and 9)
        try:
            original_start = datetime.fromisoformat(entry[8])  # start_time
            original_end = datetime.fromisoformat(entry[9])    # end_time
            original_duration = (original_end - original_start).total_seconds() / 3600
        except:
            original_start = datetime.now()
            original_end = datetime.now()
            original_duration = 0

        # Date field
        ttk.Label(form_frame, text="Date (MM/DD/YY):").grid(row=1, column=0, sticky='w', pady=5)
        date_entry = ttk.Entry(form_frame, width=15)
        date_entry.grid(row=1, column=1, sticky='w', padx=5, pady=5)
        date_entry.insert(0, original_start.strftime("%m/%d/%y"))
        
        # Entry mode selection (Start/End Time OR Decimal Hours)
        ttk.Label(form_frame, text="Entry Mode:").grid(row=2, column=0, sticky='w', pady=5)
        mode_frame = ttk.Frame(form_frame)
        mode_frame.grid(row=2, column=1, sticky='w', padx=5, pady=5)
        
        edit_entry_mode = tk.StringVar(value="decimal")  # Default to decimal
        ttk.Radiobutton(mode_frame, text="Start/End Time", variable=edit_entry_mode,
                       value="time_range").pack(side='left')
        ttk.Radiobutton(mode_frame, text="Decimal Hours", variable=edit_entry_mode,
                       value="decimal").pack(side='left', padx=10)

        # Start Time (for time_range mode)
        start_time_label = ttk.Label(form_frame, text="Start Time (HH:MM AM/PM):")
        start_time_label.grid(row=3, column=0, sticky='w', pady=5)
        start_time_entry = ttk.Entry(form_frame, width=20)
        start_time_entry.grid(row=3, column=1, sticky='w', padx=5, pady=5)
        start_time_entry.insert(0, original_start.strftime("%I:%M %p"))

        # End Time (for time_range mode)
        end_time_label = ttk.Label(form_frame, text="End Time (HH:MM AM/PM):")
        end_time_label.grid(row=4, column=0, sticky='w', pady=5)
        end_time_entry = ttk.Entry(form_frame, width=20)
        end_time_entry.grid(row=4, column=1, sticky='w', padx=5, pady=5)
        end_time_entry.insert(0, original_end.strftime("%I:%M %p"))
        
        # Decimal Hours (for decimal mode) - hidden by default
        hours_label = ttk.Label(form_frame, text="Duration (hours):")
        hours_label.grid(row=5, column=0, sticky='w', pady=5)
        hours_entry = ttk.Entry(form_frame, width=15)
        hours_entry.grid(row=5, column=1, sticky='w', padx=5, pady=5)
        hours_entry.insert(0, f"{original_duration:.2f}")
        
        hours_help = ttk.Label(form_frame, text="(e.g., 1.5, 2.25, 0.75)", 
                 font=('Arial', 8), foreground='gray')
        hours_help.grid(row=6, column=1, sticky='w', padx=5)
        
        # Toggle function
        def toggle_edit_mode():
            mode = edit_entry_mode.get()
            if mode == "time_range":
                # Show start/end time
                start_time_entry.grid()
                end_time_entry.grid()
                # Hide decimal
                hours_entry.grid_remove()
                hours_label.grid_remove()
                hours_help.grid_remove()
            else:  # decimal
                # Hide start/end time
                start_time_entry.grid_remove()
                end_time_entry.grid_remove()
                # Show decimal
                hours_entry.grid()
                hours_label.grid()
                hours_help.grid()
        
        # Bind toggle
        mode_frame.winfo_children()[0].configure(command=toggle_edit_mode)
        mode_frame.winfo_children()[1].configure(command=toggle_edit_mode)
        
        # Start in decimal mode
        toggle_edit_mode()

        # Description (column 12)
        ttk.Label(form_frame, text="Description:").grid(row=7, column=0, sticky='nw', pady=5)
        desc_text = tk.Text(form_frame, height=3, width=30)
        desc_text.grid(row=7, column=1, sticky='ew', padx=5, pady=5)
        desc_text.insert('1.0', entry[12] if entry[12] else "")  # description

        form_frame.columnconfigure(1, weight=1)

        # Buttons
        button_frame = ttk.Frame(edit_window)
        button_frame.pack(fill='x', padx=15, pady=(0, 15))

        def save_changes():
            try:
                # Get date
                date_str = date_entry.get().strip()
                date_obj = datetime.strptime(date_str, "%m/%d/%y")
                
                mode = edit_entry_mode.get()
                
                if mode == "time_range":
                    # Start/End Time mode
                    start_str = start_time_entry.get().strip()
                    end_str = end_time_entry.get().strip()
                    
                    if not start_str or not end_str:
                        messagebox.showerror("Error", "Please enter both start and end times")
                        return
                    
                    # Parse times
                    start_time = datetime.strptime(f"{date_str} {start_str}", "%m/%d/%y %I:%M %p")
                    end_time = datetime.strptime(f"{date_str} {end_str}", "%m/%d/%y %I:%M %p")
                    
                    if end_time <= start_time:
                        messagebox.showerror("Error", "End time must be after start time")
                        return
                    
                    duration_hours = (end_time - start_time).total_seconds() / 3600
                    
                else:  # decimal mode
                    # Decimal Hours mode
                    duration_str = hours_entry.get().strip()
                    
                    if not duration_str:
                        messagebox.showerror("Error", "Please enter duration in hours")
                        return
                    
                    duration_hours = float(duration_str)
                    
                    if duration_hours <= 0:
                        messagebox.showerror("Error", "Duration must be greater than 0")
                        return
                    
                    if duration_hours > 24:
                        messagebox.showerror("Error", "Duration cannot exceed 24 hours")
                        return
                    
                    # Use 9 AM as default start time for decimal mode
                    start_time = datetime.strptime(f"{date_str} 09:00 AM", "%m/%d/%y %I:%M %p")
                    end_time = start_time + timedelta(hours=duration_hours)
                
                # Get description
                description = desc_text.get('1.0', 'end').strip()

                # Update the entry
                self.time_entry_model.update(entry_id, start_time, end_time, description)
                self.refresh_time_entries()
                close_callback()
                messagebox.showinfo("Success", 
                                  f"Time entry updated successfully\n\n" +
                                  f"Duration: {duration_hours:.2f} hours\n\n" +
                                  f"Click REFRESH to update the invoice data below.")

            except ValueError as e:
                messagebox.showerror("Error", f"Invalid input: {str(e)}")

        ttk.Button(button_frame, text="Save Changes", command=save_changes).pack(side='right', padx=5)
        ttk.Button(button_frame, text="Cancel", command=close_callback).pack(side='right')

    def delete_time_entry(self):
        selection = self.entries_tree.selection()
        if not selection:
            messagebox.showerror("Error", "Please select a time entry to delete")
            return
        
        # Get the selected item
        selected_item = selection[0]
        item_tags = self.entries_tree.item(selected_item)['tags']
        
        # Only allow deletion of actual entries, not group nodes
        if 'entry' not in item_tags:
            messagebox.showerror("Error", "Please select an individual time entry (not a group)")
            return
        
        # Extract entry ID from tags
        entry_id = None
        for tag in item_tags:
            if tag.startswith('entry_id_'):
                entry_id = int(tag.replace('entry_id_', ''))
                break
        
        if not entry_id:
            messagebox.showerror("Error", "Could not find entry ID")
            return

        if messagebox.askyesno("Confirm", "Delete this time entry?"):
            self.time_entry_model.delete(entry_id)
            self.refresh_time_entries()
            messagebox.showinfo("Success", "Time entry deleted successfully")

    def export_time_entries_to_excel(self):
        """Export time entries to Excel file, respecting current filter or selection"""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment
        except ImportError as e:
            messagebox.showerror("Error",
                                 f"openpyxl library not installed.\n\nError: {e}\n\nInstall with: pip install openpyxl")
            return

        try:
            # Check if user has selected specific entries
            selected_items = self.entries_tree.selection()


            if selected_items:
                # Export only selected entries
                selected_ids = []
                for item in selected_items:
                    item_data = self.entries_tree.item(item)
                    tags = item_data['tags']
                    # Entry ID is stored in tags like 'entry_id_123'
                    if tags:
                        for tag in tags:
                            if tag.startswith('entry_id_'):
                                try:
                                    entry_id = int(tag.replace('entry_id_', ''))
                                    selected_ids.append(entry_id)
                                    break
                                except (ValueError, TypeError):
                                    pass

                if not selected_ids:
                    messagebox.showinfo("No Selection", "Please select time entries (not group headers) to export.")
                    return

                with self.db.get_connection() as conn:
                    cursor = conn.cursor()
                    placeholders = ','.join('?' * len(selected_ids))
                    cursor.execute(f'''
                        SELECT te.id, te.client_name, te.project_name, te.task_name,
                               te.start_time, te.end_time, te.duration_minutes, te.description,
                               te.is_billed, te.invoice_number, te.date
                        FROM time_entries te
                        WHERE te.id IN ({placeholders})
                        ORDER BY te.date DESC, te.start_time DESC
                    ''', selected_ids)
                    entries = cursor.fetchall()

                export_scope = "Selected"
            else:
                filter_val = self.time_entries_filter_var.get() if hasattr(self,
                                                                           'time_entries_filter_var') else 'unbilled'

                where_clause = ""
                if filter_val == "unbilled":
                    where_clause = "WHERE te.is_billed = 0"
                elif filter_val == "billed":
                    where_clause = "WHERE te.is_billed = 1"

                with self.db.get_connection() as conn:
                    cursor = conn.cursor()
                    cursor.execute(f'''
                        SELECT te.id, te.client_name, te.project_name, te.task_name,
                               te.start_time, te.end_time, te.duration_minutes, te.description,
                               te.is_billed, te.invoice_number, te.date
                        FROM time_entries te
                        {where_clause}
                        ORDER BY te.date DESC, te.start_time DESC
                    ''')
                    entries = cursor.fetchall()

                export_scope = filter_val.capitalize()

            if not entries:
                messagebox.showinfo("No Data", "No time entries found to export.")
                return

            filename = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[("Excel files", "*.xlsx")],
                initialfile=f"TimeEntries_{export_scope}_{datetime.now().strftime('%Y%m%d')}.xlsx"
            )

            if not filename:
                return

            wb = Workbook()
            ws = wb.active
            ws.title = f"Time Entries ({export_scope})"

            headers = ['Date', 'Client', 'Project', 'Task', 'Start Time', 'End Time', 'Duration (hrs)', 'Description',
                       'Billed', 'Invoice #']
            ws.append(headers)

            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center")

            for entry in entries:
                entry_id, client, project, task, start_time, end_time, duration_mins, description, is_billed, invoice_num, date = entry

                try:
                    start_dt = datetime.fromisoformat(start_time)
                    start_display = start_dt.strftime("%I:%M %p")
                except:
                    start_display = start_time

                try:
                    end_dt = datetime.fromisoformat(end_time)
                    end_display = end_dt.strftime("%I:%M %p")
                except:
                    end_display = end_time

                duration_hours = (duration_mins / 60.0) if duration_mins else 0
                billed_status = "Yes" if is_billed else "No"

                ws.append([
                    date,
                    client,
                    project,
                    task,
                    start_display,
                    end_display,
                    round(duration_hours, 2),
                    description or "",
                    billed_status,
                    invoice_num or ""
                ])

            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width

            wb.save(filename)
            messagebox.showinfo("Success", f"Exported {len(entries)} time entries to:\n\n{filename}")

            if messagebox.askyesno("Open File?", "Would you like to open the exported file now?"):
                import os
                os.startfile(filename)

        except Exception as e:
            messagebox.showerror("Export Error", f"Failed to export time entries:\n\n{e}\n\n{type(e).__name__}")
            import traceback
            traceback.print_exc()


    def invoice_selected_entries(self):
        """Generate invoice from selected time entries"""
        selection = self.entries_tree.selection()
        if not selection:
            messagebox.showerror("Error", "Please select one or more time entries to invoice")
            return

        # Get selected entry IDs and check if any are already billed
        entry_ids = []
        billed_entries = []
        client_id = None
        client_name = None

        for item in selection:
            item_tags = self.entries_tree.item(item)['tags']
            
            # Skip non-entry items (groups)
            if 'entry' not in item_tags:
                continue
            
            # Extract entry ID from tags
            entry_id = None
            for tag in item_tags:
                if tag.startswith('entry_id_'):
                    entry_id = int(tag.replace('entry_id_', ''))
                    break
            
            if not entry_id:
                continue
            
            # Get entry details from database
            all_entries = self.time_entry_model.get_all()
            entry_data = None
            for e in all_entries:
                if e[0] == entry_id:
                    entry_data = e
                    break
            
            if not entry_data:
                continue
            
            entry_client = entry_data[1]
            entry_task = entry_data[3]

            # Check if already billed
            if '[BILLED]' in entry_task:
                billed_entries.append(entry_id)
                continue

            entry_ids.append(entry_id)

            # Get client_id from first entry
            if client_id is None:
                client_name = entry_client
                # Find client ID
                clients = self.client_model.get_all()
                for client in clients:
                    if client[1] == entry_client:
                        client_id = client[0]
                        break

        # Warn about billed entries
        if billed_entries:
            messagebox.showwarning("Already Billed",
                                   f"{len(billed_entries)} selected entry(ies) already billed and will be skipped.")

        if not entry_ids:
            messagebox.showerror("Error", "No unbilled entries selected")
            return

        if not client_id:
            messagebox.showerror("Error", "Could not determine client for selected entries")
            return

        # Verify all entries are from same client
        for item in selection:
            if self.entries_tree.item(item)['values'][1] != client_name:
                messagebox.showerror("Error",
                                     "All selected entries must be from the same client.\n"
                                     "Please select entries from one client only.")
                return

        # Confirm before generating invoice
        if not messagebox.askyesno("Confirm Invoice",
                                   f"Generate invoice for {len(entry_ids)} time entry(ies) from {client_name}?"):
            return

        # Generate invoice from selected entries
        self.generate_invoice_from_entries(client_id, client_name, entry_ids)

    def generate_invoice_from_entries(self, client_id, client_name, entry_ids):
        """Generate invoice data from specific entry IDs"""
        import sqlite3

        # Get entry details
        conn = self.db.conn
        conn.row_factory = sqlite3.Row
        cursor = conn.cursor()

        placeholders = ','.join(['?' for _ in entry_ids])
        cursor.execute(f'''
            SELECT * FROM invoice_view
            WHERE entry_id IN ({placeholders})
            AND (is_billed = 0 OR is_billed IS NULL)
        ''', entry_ids)

        entries = cursor.fetchall()
        conn.row_factory = None

        if not entries:
            messagebox.showinfo("No Entries", "Selected entries are already billed or not found.")
            return

        self.pending_entry_ids = [row['entry_id'] for row in entries]
        invoice_items = []

        # Group by task
        tasks = {}
        for row in entries:
            key = f"{row['project_name']} - {row['task_name']}"
            if key not in tasks:
                # Check if it's a lump sum task
                if row['task_lump_sum']:
                    tasks[key] = {
                        'minutes': 0,
                        'rate': 0,
                        'is_lump_sum': True,
                        'lump_sum_amount': row['task_lump_amount']
                    }
                elif row['project_lump_sum']:
                    tasks[key] = {
                        'minutes': 0,
                        'rate': 0,
                        'is_lump_sum': True,
                        'lump_sum_amount': row['project_lump_amount']
                    }
                else:
                    tasks[key] = {
                        'minutes': 0,
                        'rate': row['task_rate'] or row['project_rate'],
                        'is_lump_sum': False,
                        'lump_sum_amount': 0
                    }

            tasks[key]['minutes'] += row['duration_minutes'] or 0

        # Build invoice items
        for task_name, data in tasks.items():
            hours = data['minutes'] / 60.0

            if data['is_lump_sum']:
                invoice_items.append({
                    'description': task_name,
                    'quantity': '1',
                    'rate': f"${data['lump_sum_amount']:.2f}",
                    'amount': data['lump_sum_amount']
                })
            else:
                invoice_items.append({
                    'description': task_name,
                    'quantity': f"{hours:.2f} hrs",
                    'rate': f"${data['rate']:.2f}/hr",
                    'amount': hours * data['rate']
                })

        # Get earliest and latest dates from selected entries
        start_date = None
        end_date = None
        for row in entries:
            entry_date = datetime.fromisoformat(row['start_time']).date()
            if start_date is None or entry_date < start_date:
                start_date = entry_date
            if end_date is None or entry_date > end_date:
                end_date = entry_date

        # Create invoice data
        self.current_invoice_data = {
            'client_id': client_id,
            'start_date': datetime.combine(start_date, datetime.min.time()),
            'end_date': datetime.combine(end_date, datetime.max.time()),
            'items': invoice_items,
            'total': sum(item['amount'] for item in invoice_items)
        }

        # Switch to Invoices tab and display preview
        self.notebook.select(6)  # Invoices tab is index 6
        self.display_invoice_preview()

        messagebox.showinfo("Invoice Generated",
                            f"Invoice preview ready for {client_name}\n"
                            f"Total: ${self.current_invoice_data['total']:.2f}\n\n"
                            "Review and click 'Save as PDF' to finalize.")

    # Company info methods
    def browse_logo(self):
        filename = filedialog.askopenfilename(
            title="Select Logo Image",
            filetypes=[("Image files", "*.png *.jpg *.jpeg *.gif *.bmp")]
        )
        if filename:
            self.logo_path_var.set(filename)

    def save_company_info(self):
        name = self.company_name_entry.get().strip()
        address = self.company_address_text.get("1.0", tk.END).strip()
        phone = self.company_phone_entry.get().strip()
        email = self.company_email_entry.get().strip()
        logo_path = self.logo_path_var.get()
        website = self.company_website_entry.get().strip()
        payment_terms = self.company_payment_terms_entry.get().strip()
        thank_you_message = self.company_thank_you_entry.get().strip()

        if not name:
            messagebox.showerror("Error", "Company name is required")
            return

        # Save to database with new fields
        try:
            with self.db.get_connection() as conn:
                cursor = conn.cursor()
                cursor.execute('''
                    INSERT OR REPLACE INTO company_info 
                    (id, name, address, phone, email, logo_path, website, payment_terms, thank_you_message)
                    VALUES (1, ?, ?, ?, ?, ?, ?, ?, ?)
                ''', (name, address, phone, email, logo_path, website, payment_terms, thank_you_message))
                conn.commit()
            messagebox.showinfo("Success", "Company information saved successfully")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to save company info: {str(e)}")

    def load_company_info(self):
        company = self.company_model.get()
        if company:
            self.company_name_entry.delete(0, tk.END)
            self.company_name_entry.insert(0, company[1])

            self.company_address_text.delete("1.0", tk.END)
            self.company_address_text.insert("1.0", company[2] or "")

            self.company_phone_entry.delete(0, tk.END)
            self.company_phone_entry.insert(0, company[3] or "")

            self.company_email_entry.delete(0, tk.END)
            self.company_email_entry.insert(0, company[4] or "")

            if len(company) > 5 and company[5]:
                self.logo_path_var.set(company[5])
            
            # Load new fields (website, payment_terms, thank_you_message)
            if len(company) > 6 and company[6]:
                self.company_website_entry.delete(0, tk.END)
                self.company_website_entry.insert(0, company[6])
            
            if len(company) > 7 and company[7]:
                self.company_payment_terms_entry.delete(0, tk.END)
                self.company_payment_terms_entry.insert(0, company[7])
            else:
                # Default value if not in database yet
                self.company_payment_terms_entry.delete(0, tk.END)
                self.company_payment_terms_entry.insert(0, "Payment is due within 30 days")
            
            if len(company) > 8 and company[8]:
                self.company_thank_you_entry.delete(0, tk.END)
                self.company_thank_you_entry.insert(0, company[8])
            else:
                # Default value if not in database yet
                self.company_thank_you_entry.delete(0, tk.END)
                self.company_thank_you_entry.insert(0, "Thank you for your business!")

    # Invoice methods
    def generate_invoice(self):
        client_text = self.invoice_client_combo.get()

        if not client_text:
            messagebox.showerror("Error", "Please select a client")
            return

        # Get client ID
        clients = self.client_model.get_all()
        client_id = None
        for client in clients:
            if client[1] == client_text:
                client_id = client[0]
                break

        if not client_id:
            messagebox.showerror("Error", "Invalid client selected")
            return

        try:
            start_date = datetime.strptime(self.invoice_start_date.get(), "%m/%d/%y")
            end_date = datetime.strptime(self.invoice_end_date.get(), "%m/%d/%y")
        except ValueError:
            messagebox.showerror("Error", "Invalid date format. Use MM/DD/YY")
            return

        # Generate invoice data
        self.current_invoice_data = self.generate_invoice_data(client_id, start_date, end_date)
        self.display_invoice_preview()

    def generate_invoice_data(self, client_id, start_date, end_date):
        conn = self.db.get_connection()
        conn.row_factory = sqlite3.Row
        cursor = conn.cursor()

        cursor.execute('''
                       SELECT *
                       FROM invoice_view
                       WHERE client_id = ?
                         AND DATE (start_time) BETWEEN ?
                         AND ?
                         AND (is_billed = 0
                          OR is_billed IS NULL)
                       ''', (client_id, start_date.strftime("%Y-%m-%d"), end_date.strftime("%Y-%m-%d")))

        entries = cursor.fetchall()
        conn.row_factory = None

        if not entries:
            messagebox.showinfo("No Unbilled Hours", "No unbilled hours found for the selected client and date range.")
            return None

        self.pending_entry_ids = [row['entry_id'] for row in entries]
        invoice_items = []

        # Simple task grouping
        tasks = {}
        for row in entries:
            key = f"{row['project_name']} - {row['task_name']}"
            if key not in tasks:
                tasks[key] = {'minutes': 0, 'rate': row['task_rate'] or row['project_rate']}
            tasks[key]['minutes'] += row['duration_minutes'] or 0

        for task_name, data in tasks.items():
            hours = data['minutes'] / 60.0
            invoice_items.append({
                'description': task_name,
                'quantity': f"{hours:.2f} hrs",
                'rate': f"${data['rate']:.2f}/hr",
                'amount': hours * data['rate']
            })

        return {
            'client_id': client_id,
            'start_date': start_date,
            'end_date': end_date,
            'items': invoice_items,
            'total': sum(item['amount'] for item in invoice_items)
        }

    def display_invoice_preview(self):
        # Clear existing items
        for item in self.invoice_tree.get_children():
            self.invoice_tree.delete(item)

        # Check if invoice data exists
        if not self.current_invoice_data:
            self.invoice_total_label.config(text="Total: $0.00")
            return

        # Add invoice items
        for item in self.current_invoice_data['items']:
            self.invoice_tree.insert('', 'end', values=(
                item['description'],
                item['quantity'],
                item['rate'],
                f"${item['amount']:.2f}"
            ))

        # Update total
        self.invoice_total_label.config(text=f"Total: ${self.current_invoice_data['total']:.2f}")

    def save_invoice_pdf(self):
        if not hasattr(self, 'current_invoice_data'):
            messagebox.showerror("Error", "Please generate an invoice first")
            return

        if not self.current_invoice_data:
            messagebox.showerror("Error", "No invoice data available")
            return

        # Generate invoice number based on date
        invoice_date = datetime.now()
        invoice_number = f"INV-{invoice_date.strftime('%Y%m%d-%H%M%S')}"

        filename = filedialog.asksaveasfilename(
            defaultextension=".pdf",
            filetypes=[("PDF files", "*.pdf")],
            initialfile=f"Invoice_{invoice_number}.pdf"
        )

        if filename:
            try:
                from invoice_generator import InvoiceGenerator
                generator = InvoiceGenerator(self.db)
                generator.generate_pdf(self.current_invoice_data, filename, invoice_number)

                # Mark time entries as billed
                if hasattr(self, 'pending_entry_ids') and self.pending_entry_ids:
                    self.db.mark_entries_billed(self.pending_entry_ids, invoice_number)

                    # Save to billing history
                    self.db.save_billing_history(self.current_invoice_data, invoice_number, filename)

                    # Clear pending entries
                    self.pending_entry_ids = []

                    messagebox.showinfo("Success",
                                        f"Invoice saved as {filename}\n\n"
                                        f"Time entries have been marked as billed and will not appear in future invoices.")
                else:
                    messagebox.showinfo("Success", f"Invoice saved as {filename}")

                # Clear current invoice data
                self.current_invoice_data = None

                # Clear the preview
                for item in self.invoice_tree.get_children():
                    self.invoice_tree.delete(item)
                self.invoice_total_label.config(text="Total: $0.00")

                # Refresh time entries to show billed status
                self.refresh_time_entries()

            except Exception as e:
                messagebox.showerror("Error", f"Failed to save invoice: {str(e)}")

    def edit_invoice_item(self):
        selection = self.invoice_tree.selection()
        if not selection:
            messagebox.showerror("Error", "Please select an item to edit")
            return

        item_values = self.invoice_tree.item(selection[0])['values']
        self.open_edit_invoice_item_dialog(selection[0], item_values)

    def open_edit_invoice_item_dialog(self, item_id, item_values):
        edit_window = tk.Toplevel(self.root)
        edit_window.title("Edit Invoice Item")
        self.center_dialog(edit_window, 400, 200)

        form_frame = ttk.Frame(edit_window)
        form_frame.pack(fill='both', expand=True, padx=10, pady=10)

        ttk.Label(form_frame, text="Description:").grid(row=0, column=0, sticky='w', pady=2)
        desc_entry = ttk.Entry(form_frame, width=40)
        desc_entry.grid(row=0, column=1, sticky='ew', padx=5, pady=2)
        desc_entry.insert(0, item_values[0])

        ttk.Label(form_frame, text="Quantity:").grid(row=1, column=0, sticky='w', pady=2)
        qty_entry = ttk.Entry(form_frame)
        qty_entry.grid(row=1, column=1, sticky='ew', padx=5, pady=2)
        qty_entry.insert(0, item_values[1])

        ttk.Label(form_frame, text="Rate:").grid(row=2, column=0, sticky='w', pady=2)
        rate_entry = ttk.Entry(form_frame)
        rate_entry.grid(row=2, column=1, sticky='ew', padx=5, pady=2)
        rate_entry.insert(0, item_values[2])

        form_frame.columnconfigure(1, weight=1)

        button_frame = ttk.Frame(edit_window)
        button_frame.pack(fill='x', padx=10, pady=10)

        def save_changes():
            # Update the tree item
            self.invoice_tree.item(item_id, values=(
                desc_entry.get(),
                qty_entry.get(),
                rate_entry.get(),
                item_values[3]
            ))

            # Recalculate if needed
            if hasattr(self, 'current_invoice_data'):
                for item in self.current_invoice_data['items']:
                    if item['description'] == item_values[0]:
                        item['description'] = desc_entry.get()
                        item['quantity'] = qty_entry.get()
                        item['rate'] = rate_entry.get()
                        break

            edit_window.destroy()

        ttk.Button(button_frame, text="Save", command=save_changes).pack(side='right', padx=5)
        ttk.Button(button_frame, text="Cancel", command=edit_window.destroy).pack(side='right')

    def remove_invoice_item(self):
        selection = self.invoice_tree.selection()
        if not selection:
            messagebox.showerror("Error", "Please select an item to remove")
            return

        # Get item description before removing
        item_values = self.invoice_tree.item(selection[0])['values']

        # Remove from tree
        self.invoice_tree.delete(selection[0])

        # Update invoice data if it exists
        if hasattr(self, 'current_invoice_data'):
            self.current_invoice_data['items'] = [
                item for item in self.current_invoice_data['items']
                if item['description'] != item_values[0]
            ]

            # Recalculate total
            self.current_invoice_data['total'] = sum(
                item['amount'] for item in self.current_invoice_data['items']
            )
            self.invoice_total_label.config(
                text=f"Total: ${self.current_invoice_data['total']:.2f}"
            )

    # Refresh methods
    def refresh_clients(self):
        # Clear tree
        for item in self.client_tree.get_children():
            self.client_tree.delete(item)

        # Add clients
        clients = self.client_model.get_all()
        for client in clients:
            self.client_tree.insert('', 'end', values=(
                client[0],  # ID
                client[1],  # Name
                client[2] or "",  # Company
                client[3] or "",  # Email
                client[4] or ""  # Phone
            ))

    def refresh_projects(self):
        # Clear tree
        for item in self.project_tree.get_children():
            self.project_tree.delete(item)

        # Add projects
        projects = self.project_model.get_all()
        for project in projects:
            # Project structure from models.py get_all():
            # 0: id, 1: client_id, 2: name, 3: description,
            # 4: hourly_rate, 5: is_lump_sum, 6: lump_sum_amount,
            # 7: created_at, 8: updated_at, 9: client_name (from JOIN)

            billing_type = "Lump Sum" if project[5] else "Hourly"
            rate = f"${project[6]:.2f}" if project[5] else f"${project[4]:.2f}/hr"
            client_name = project[9] if len(project) > 9 else "Unknown Client"

            self.project_tree.insert('', 'end', values=(
                project[0],  # ID
                client_name,  # Client name (was project[7], should be project[9])
                project[2],  # Project name
                billing_type,
                rate
            ))

    def refresh_tasks(self):
        # Save expansion state before clearing
        expanded_items = self.save_tree_state(self.task_tree)
        
        # Clear tree
        for item in self.task_tree.get_children():
            self.task_tree.delete(item)

        # Get all tasks and organize hierarchically
        tasks = self.task_model.get_all()
        global_tasks = self.task_model.get_global_tasks()
        
        # Build hierarchy: {client_name: {project_name: [tasks]}}
        hierarchy = {}
        
        for task in tasks:
            client_name = task[9] if len(task) > 9 and task[9] else "Unknown Client"
            project_name = task[8] if len(task) > 8 and task[8] else "Unknown Project"
            
            if client_name not in hierarchy:
                hierarchy[client_name] = {}
            if project_name not in hierarchy[client_name]:
                hierarchy[client_name][project_name] = []
            
            hierarchy[client_name][project_name].append(task)
        
        # Display hierarchy
        for client_name in sorted(hierarchy.keys()):
            # Insert client node
            client_id = self.task_tree.insert('', 'end',
                text=f"📁 {client_name}",
                values=('Client', '', '', '', ''),
                tags=('client', 'client_row'))
            
            for project_name in sorted(hierarchy[client_name].keys()):
                # Insert project node
                project_id = self.task_tree.insert(client_id, 'end',
                    text=f"  📂 {project_name}",
                    values=('Project', '', '', '', ''),
                    tags=('project', 'project_row'))
                
                # Insert tasks under project
                for task in hierarchy[client_name][project_name]:
                    billing_type = "Lump Sum" if task[5] else "Hourly"
                    rate = f"${task[6]:.2f}" if task[5] else f"${task[4]:.2f}/hr"
                    
                    self.task_tree.insert(project_id, 'end',
                        text=f"    ⚙️ {task[2]}",
                        values=('Task', task[0], '', billing_type, rate),
                        tags=('task', 'task_row', f'task_id_{task[0]}'))
        
        # Add global tasks section if any exist
        if global_tasks:
            global_node = self.task_tree.insert('', 'end',
                text="📁 [GLOBAL TASKS]",
                values=('Global', '', '', '', ''),
                tags=('global',))
            
            for task in global_tasks:
                billing_type = "Lump Sum" if task[5] else "Hourly"
                rate = f"${task[6]:.2f}" if task[5] else f"${task[4]:.2f}/hr"
                
                self.task_tree.insert(global_node, 'end',
                    text=f"  ⚙️ {task[2]}",
                    values=('Global Task', task[0], '', billing_type, rate),
                    tags=('task', f'task_id_{task[0]}'))
        
        # Configure tag colors
        self.task_tree.tag_configure('client', font=('Arial', 10, 'bold'))
        self.task_tree.tag_configure('project', font=('Arial', 9, 'bold'))
        self.task_tree.tag_configure('task', font=('Arial', 9))
        self.task_tree.tag_configure('global', font=('Arial', 10, 'bold'), foreground='#10b981')
        
        # Restore expansion state (expand all by default)
        self.restore_tree_state(self.task_tree, expanded_items, expand_all=True)


    def refresh_time_entries(self):
        # Save expansion state before clearing
        expanded_items = self.save_tree_state(self.entries_tree)
        
        # Clear tree
        for item in self.entries_tree.get_children():
            self.entries_tree.delete(item)

        # Get filter value
        filter_val = self.time_entries_filter_var.get() if hasattr(self, 'time_entries_filter_var') else 'unbilled'
        
        # Build WHERE clause
        where_clause = ""
        if filter_val == "unbilled":
            where_clause = "WHERE te.is_billed = 0"
        elif filter_val == "billed":
            where_clause = "WHERE te.is_billed = 1"
        # else "all" - no filter
        
        # Get time entries with filter
        with self.db.get_connection() as conn:
            cursor = conn.cursor()
            cursor.execute(f'''
                SELECT te.id, te.client_name, te.project_name, te.task_name,
                       te.start_time, te.end_time, te.duration_minutes, te.description,
                       te.is_billed, te.invoice_number
                FROM time_entries te
                {where_clause}
                ORDER BY te.client_name, te.project_name, te.task_name, te.start_time DESC
            ''')
            entries = cursor.fetchall()
        
        if not entries:
            return
        
        # Group entries by Client > Project > Task
        hierarchy = {}  # {client_name: {project_name: {task_name: [entries]}}}
        
        for entry in entries:
            client_name = entry[1]
            project_name = entry[2]
            task_name = entry[3]
            
            if client_name not in hierarchy:
                hierarchy[client_name] = {}
            if project_name not in hierarchy[client_name]:
                hierarchy[client_name][project_name] = {}
            if task_name not in hierarchy[client_name][project_name]:
                hierarchy[client_name][project_name][task_name] = []
            
            hierarchy[client_name][project_name][task_name].append(entry)
        
        # Build the tree structure
        for client_name in sorted(hierarchy.keys()):
            # Calculate client total
            client_total_minutes = 0
            for project_name in hierarchy[client_name]:
                for task_name in hierarchy[client_name][project_name]:
                    for entry in hierarchy[client_name][project_name][task_name]:
                        client_total_minutes += entry[6] if entry[6] else 0
            
            client_total_hours = client_total_minutes / 60.0
            
            # Insert client node
            client_id = self.entries_tree.insert('', 'end', 
                text=f"📁 {client_name}",
                values=('Client', '', '', f"{client_total_hours:.2f} hrs", ''),
                tags=('client', 'client_row'))
            
            # Add projects under client
            for project_name in sorted(hierarchy[client_name].keys()):
                # Calculate project total
                project_total_minutes = 0
                for task_name in hierarchy[client_name][project_name]:
                    for entry in hierarchy[client_name][project_name][task_name]:
                        project_total_minutes += entry[6] if entry[6] else 0
                
                project_total_hours = project_total_minutes / 60.0
                
                # Insert project node
                project_id = self.entries_tree.insert(client_id, 'end',
                    text=f"  📂 {project_name}",
                    values=('Project', '', '', f"{project_total_hours:.2f} hrs", ''),
                    tags=('project', 'project_row'))
                
                # Add tasks under project
                for task_name in sorted(hierarchy[client_name][project_name].keys()):
                    task_entries = hierarchy[client_name][project_name][task_name]
                    
                    # Calculate task total
                    task_total_minutes = sum(e[6] if e[6] else 0 for e in task_entries)
                    task_total_hours = task_total_minutes / 60.0
                    
                    # Check if any entry is billed
                    has_billed = any(e[8] for e in task_entries)
                    billed_indicator = " [BILLED]" if has_billed else ""
                    
                    # Insert task node
                    task_id = self.entries_tree.insert(project_id, 'end',
                        text=f"    📋 {task_name}{billed_indicator}",
                        values=('Task', f"{len(task_entries)} entries", '', f"{task_total_hours:.2f} hrs", ''),
                        tags=('task', 'task_row'))
                    
                    # Add individual entries under task
                    for entry in task_entries:
                        duration_minutes = entry[6] if entry[6] else 0
                        duration_hours = duration_minutes / 60.0
                        
                        # Format start time
                        start_time = entry[4]
                        try:
                            dt = datetime.fromisoformat(start_time)
                            start_display = dt.strftime("%m/%d/%y %I:%M %p")
                        except:
                            start_display = start_time
                        
                        # Billed indicator for individual entry
                        is_billed = entry[8]
                        entry_billed = " [BILLED]" if is_billed else ""
                        
                        # Store entry ID in tags for later retrieval
                        # Calculate row number for alternating colors
                        entry_index = task_entries.index(entry)
                        row_tag = 'oddrow' if entry_index % 2 == 0 else 'evenrow'
                        
                        entry_id = self.entries_tree.insert(task_id, 'end',
                            text=f"      ⏱️ Entry",
                            values=('Entry' + entry_billed, '', start_display, f"{duration_hours:.2f} hrs", entry[7] or ''),
                            tags=('entry', f'entry_id_{entry[0]}', 'entry_row'))
        
        # Configure tag colors
        self.entries_tree.tag_configure('client', font=('Arial', 10, 'bold'))
        self.entries_tree.tag_configure('project', font=('Arial', 9, 'bold'))
        self.entries_tree.tag_configure('task', font=('Arial', 9))
        self.entries_tree.tag_configure('entry', font=('Arial', 8))
        
        # Restore expansion state (expand all by default)
        self.restore_tree_state(self.entries_tree, expanded_items, expand_all=True)

    def refresh_combos(self):
        # Refresh timer combos
        clients = self.client_model.get_all()
        client_names = [c[1] for c in clients]
        self.timer_client_combo['values'] = client_names

        # Refresh project combo
        self.project_client_combo['values'] = client_names

        # Refresh task client combo
        self.task_client_combo['values'] = client_names

        # Refresh task combo for manual entry (include global tasks)
        tasks = self.task_model.get_all()
        global_tasks = self.task_model.get_global_tasks()
        
        task_displays = []
        
        # Add global tasks first with [GLOBAL] prefix
        for task in global_tasks:
            task_displays.append(f"[GLOBAL] {task[2]}")
        
        # Add regular project tasks
        for task in tasks:
            client_name = task[9] if len(task) > 9 else "Unknown"
            project_name = task[8] if len(task) > 8 else "Unknown"
            task_displays.append(f"{client_name} - {project_name} - {task[2]}")
        
        self.manual_task_combo['values'] = task_displays

        # Refresh invoice client combo
        self.invoice_client_combo['values'] = client_names
        
        # Refresh manual entry client combo
        self.manual_client_combo['values'] = client_names

    def generate_invoice_data(self, client_id, start_date, end_date):
        # Get connection directly (not using context manager for row_factory)
        conn = self.db.conn
        conn.row_factory = sqlite3.Row
        cursor = conn.cursor()

        cursor.execute('''
                       SELECT *
                       FROM invoice_view
                       WHERE client_id = ?
                         AND DATE (start_time) BETWEEN ?
                         AND ?
                         AND (is_billed = 0
                          OR is_billed IS NULL)
                       ''', (client_id, start_date.strftime("%Y-%m-%d"), end_date.strftime("%Y-%m-%d")))

        entries = cursor.fetchall()
        conn.row_factory = None  # Reset to default

        if not entries:
            messagebox.showinfo("No Unbilled Hours", "No unbilled hours found for the selected client and date range.")
            return None

        self.pending_entry_ids = [row['entry_id'] for row in entries]
        invoice_items = []

        # Simple task grouping
        tasks = {}
        for row in entries:
            key = f"{row['project_name']} - {row['task_name']}"
            if key not in tasks:
                tasks[key] = {'minutes': 0, 'rate': row['task_rate'] or row['project_rate']}
            tasks[key]['minutes'] += row['duration_minutes'] or 0

        for task_name, data in tasks.items():
            hours = data['minutes'] / 60.0
            invoice_items.append({
                'description': task_name,
                'quantity': f"{hours:.2f} hrs",
                'rate': f"${data['rate']:.2f}/hr",
                'amount': hours * data['rate']
            })

        return {
            'client_id': client_id,
            'start_date': start_date,
            'end_date': end_date,
            'items': invoice_items,
            'total': sum(item['amount'] for item in invoice_items)
        }

    # Email Settings Methods
    def on_email_provider_select(self):
        """Auto-fill SMTP settings based on provider selection"""
        provider = self.email_provider_var.get()
        
        if provider == "gmail":
            self.smtp_server_entry.delete(0, tk.END)
            self.smtp_server_entry.insert(0, "smtp.gmail.com")
            self.smtp_port_entry.delete(0, tk.END)
            self.smtp_port_entry.insert(0, "587")
        elif provider == "outlook":
            self.smtp_server_entry.delete(0, tk.END)
            self.smtp_server_entry.insert(0, "smtp.office365.com")
            self.smtp_port_entry.delete(0, tk.END)
            self.smtp_port_entry.insert(0, "587")
        # Custom provider - let user fill manually
    
    def toggle_password_visibility(self):
        """Show/hide email password"""
        if self.show_password_var.get():
            self.email_password_entry.config(show='')
        else:
            self.email_password_entry.config(show='*')
    
    def save_email_settings(self):
        """Save email settings to database"""
        smtp_server = self.smtp_server_entry.get().strip()
        smtp_port = self.smtp_port_entry.get().strip()
        email_address = self.email_address_entry.get().strip()
        email_password = self.email_password_entry.get().strip()
        from_name = self.email_from_name_entry.get().strip()
        
        if not all([smtp_server, smtp_port, email_address, email_password]):
            messagebox.showerror("Error", "Please fill in all required fields (SMTP server, port, email, password)")
            return
        
        try:
            smtp_port = int(smtp_port)
        except ValueError:
            messagebox.showerror("Error", "Port must be a number")
            return
        
        # Save to database
        success = self.db.save_email_settings(
            smtp_server=smtp_server,
            smtp_port=smtp_port,
            email_address=email_address,
            email_password=email_password,
            from_name=from_name or None,
            send_copy_to_self=self.send_copy_to_self_var.get(),
            show_preview_before_send=self.show_preview_var.get()
        )
        
        if success:
            messagebox.showinfo("Success", "Email settings saved successfully!\n\nClick 'Test Connection' to verify.")
        else:
            messagebox.showerror("Error", "Failed to save email settings")
    
    def load_email_settings(self):
        """Load email settings from database"""
        settings = self.db.get_email_settings()
        
        if not settings:
            messagebox.showinfo("No Settings", "No email settings found. Please configure your SMTP settings.")
            return
        
        # Populate fields
        self.smtp_server_entry.delete(0, tk.END)
        self.smtp_server_entry.insert(0, settings[1])  # smtp_server
        
        self.smtp_port_entry.delete(0, tk.END)
        self.smtp_port_entry.insert(0, str(settings[2]))  # smtp_port
        
        self.email_address_entry.delete(0, tk.END)
        self.email_address_entry.insert(0, settings[3])  # email_address
        
        self.email_password_entry.delete(0, tk.END)
        self.email_password_entry.insert(0, settings[4])  # email_password
        
        if settings[5]:  # from_name
            self.email_from_name_entry.delete(0, tk.END)
            self.email_from_name_entry.insert(0, settings[5])
        
        self.send_copy_to_self_var.set(bool(settings[6]))  # send_copy_to_self
        self.show_preview_var.set(bool(settings[7]))  # show_preview_before_send
        
        messagebox.showinfo("Loaded", "Email settings loaded successfully!")
    
    def test_email_connection(self):
        """Test SMTP connection and send test email"""
        smtp_server = self.smtp_server_entry.get().strip()
        smtp_port = self.smtp_port_entry.get().strip()
        email_address = self.email_address_entry.get().strip()
        email_password = self.email_password_entry.get().strip()
        
        if not all([smtp_server, smtp_port, email_address, email_password]):
            messagebox.showerror("Error", "Please fill in all fields first")
            return
        
        try:
            smtp_port = int(smtp_port)
        except ValueError:
            messagebox.showerror("Error", "Port must be a number")
            return
        
        # Import email sender
        from email_sender import EmailSender
        
        # Create sender and test
        sender = EmailSender(smtp_server, smtp_port, email_address, email_password)
        
        # First test connection
        success, message = sender.test_connection()
        
        if success:
            # Connection worked, now send test email
            if messagebox.askyesno("Connection Successful!", 
                "SMTP connection successful! ✅\n\nWould you like to send a test email to yourself?"):
                success, message = sender.send_test_email()
                messagebox.showinfo("Test Email", message)
        else:
            messagebox.showerror("Connection Failed", message)
    
    # Email Template Methods
    def on_template_select(self, event=None):
        """Template selected from dropdown"""
        pass  # Just for binding, actual load happens on Load button
    
    def load_selected_template(self):
        """Load the selected template"""
        template_name = self.template_combo.get()
        if not template_name:
            messagebox.showwarning("No Selection", "Please select a template first")
            return
        
        # Try to load from database first
        template = self.db.get_email_template(template_name=template_name)
        
        if template:
            # Load from database
            print(f"[DEBUG] Loading template '{template_name}' from DATABASE")
            print(f"[DEBUG] Subject: {template[2][:50]}...")
            print(f"[DEBUG] Body length: {len(template[3])} chars")
            
            self.template_subject_entry.delete(0, tk.END)
            self.template_subject_entry.insert(0, template[2])  # subject
            
            self.template_body_text.delete('1.0', tk.END)
            self.template_body_text.insert('1.0', template[3])  # body
        else:
            # Load from defaults
            print(f"[DEBUG] Template '{template_name}' NOT in database, loading from DEFAULTS")
            from email_sender import EmailTemplate
            default_template = EmailTemplate.get_template(template_name)
            
            if default_template:
                self.template_subject_entry.delete(0, tk.END)
                self.template_subject_entry.insert(0, default_template['subject'])
                
                self.template_body_text.delete('1.0', tk.END)
                self.template_body_text.insert('1.0', default_template['body'])
            else:
                messagebox.showerror("Error", f"Template '{template_name}' not found")
                return
        
        self.update_template_preview()
    
    def reset_template_to_default(self):
        """Reset current template to default version"""
        template_name = self.template_combo.get()
        if not template_name:
            messagebox.showwarning("No Selection", "Please select a template first")
            return
        
        from email_sender import EmailTemplate
        default_template = EmailTemplate.get_template(template_name)
        
        if default_template:
            self.template_subject_entry.delete(0, tk.END)
            self.template_subject_entry.insert(0, default_template['subject'])
            
            self.template_body_text.delete('1.0', tk.END)
            self.template_body_text.insert('1.0', default_template['body'])
            
            self.update_template_preview()
            messagebox.showinfo("Reset", f"Template '{template_name}' reset to default")
        else:
            messagebox.showerror("Error", f"Default template '{template_name}' not found")
    
    def insert_variable(self, variable):
        """Insert a variable at cursor position in body text"""
        self.template_body_text.insert(tk.INSERT, variable)
        self.template_body_text.focus_set()
    
    def update_template_preview(self):
        """Update the preview pane with rendered template"""
        from email_sender import EmailTemplate
        import html
        
        # Get current template text
        subject = self.template_subject_entry.get()
        body = self.template_body_text.get('1.0', tk.END)
        
        # Sample data
        sample_data = {
            'client_name': 'John Smith',
            'client_company': 'Smith & Associates',
            'client_email': 'john@company.com',
            'invoice_number': 'INV-20260129-123456',
            'invoice_date': 'January 29, 2026',
            'invoice_total': '$1,234.56',
            'payment_terms': 'Payment due within 30 days',
            'due_date': 'February 28, 2026',
            'date_range': 'January 1-29, 2026',
            'company_name': 'Your Company Name',
            'company_email': 'you@company.com',
            'company_phone': '(555) 123-4567',
            'company_website': 'www.yourcompany.com'
        }
        
        # Render template
        rendered_subject = EmailTemplate.render_template(subject, sample_data)
        rendered_body = EmailTemplate.render_template(body, sample_data)
        
        # Strip HTML for preview (simple approach)
        import re
        preview_body = re.sub('<[^<]+?>', '', rendered_body)
        preview_body = html.unescape(preview_body)
        
        # Update preview
        self.template_preview_text.config(state='normal')
        self.template_preview_text.delete('1.0', tk.END)
        self.template_preview_text.insert('1.0', f"Subject: {rendered_subject}\n\n{preview_body}")
        self.template_preview_text.config(state='disabled')
    
    def save_current_template(self):
        """Save the current template to database"""
        print("[DEBUG] ===== SAVE_CURRENT_TEMPLATE CALLED =====")
        template_name = self.template_combo.get()
        print(f"[DEBUG] Template name: {template_name}")
        if not template_name:
            messagebox.showwarning("No Selection", "Please select a template first")
            return
        
        subject = self.template_subject_entry.get().strip()
        body = self.template_body_text.get('1.0', tk.END).strip()
        
        if not subject or not body:
            messagebox.showerror("Error", "Subject and body cannot be empty")
            return
        
        # Save to database
        success = self.db.save_email_template(template_name, subject, body, is_default=False)
        
        if success:
            # Verify by reloading from database
            saved_template = self.db.get_email_template(template_name=template_name)
            if saved_template and saved_template[2] == subject and saved_template[3] == body:
                messagebox.showinfo("Success", f"Template '{template_name}' saved successfully!")
            else:
                messagebox.showwarning("Warning", f"Template saved but verification failed. Try loading the template to confirm.")
        else:
            messagebox.showerror("Error", "Failed to save template")
    
    def send_test_template_email(self):
        """Send a test email using current template"""
        # Check if email settings exist
        settings = self.db.get_email_settings()
        if not settings:
            messagebox.showerror("No Email Settings", 
                "Please configure your email settings in the Email Settings tab first")
            return
        
        # Get current template
        subject = self.template_subject_entry.get().strip()
        body = self.template_body_text.get('1.0', tk.END).strip()
        
        if not subject or not body:
            messagebox.showerror("Error", "Please create a template first")
            return
        
        from email_sender import EmailSender, EmailTemplate
        
        # Sample data
        sample_data = {
            'client_name': 'John Smith',
            'client_company': 'Smith & Associates',
            'client_email': 'john@company.com',
            'invoice_number': 'INV-TEST-123456',
            'invoice_date': datetime.now().strftime('%B %d, %Y'),
            'invoice_total': '$1,234.56',
            'payment_terms': 'Payment due within 30 days',
            'due_date': (datetime.now() + timedelta(days=30)).strftime('%B %d, %Y'),
            'date_range': 'January 1-29, 2026',
            'company_name': settings[5] or 'Your Company',
            'company_email': settings[3],
            'company_phone': '(555) 123-4567',
            'company_website': 'www.yourcompany.com'
        }
        
        # Render template
        rendered_subject = EmailTemplate.render_template(subject, sample_data)
        rendered_body = EmailTemplate.render_template(body, sample_data)
        
        # Create sender
        sender = EmailSender(settings[1], settings[2], settings[3], settings[4])
        
        # Send test email
        success, message = sender.send_email(
            to_address=settings[3],  # Send to self
            subject=f"TEST: {rendered_subject}",
            body_html=rendered_body,
            from_name=settings[5]
        )
        
        if success:
            messagebox.showinfo("Success", f"{message}\n\nCheck your inbox!")
        else:
            messagebox.showerror("Failed", message)
    
    def refresh_all_data(self):
        self.refresh_clients()
        self.refresh_projects()
        self.refresh_tasks()
        self.refresh_time_entries()
        self.refresh_combos()
        self.load_company_info()
        self.update_daily_totals_display()
        self.refresh_email_templates()
        self.load_email_settings_silent()  # Auto-load email settings on startup  # Initialize daily totals display
        
        # Load email templates list if tab exists
        if hasattr(self, 'template_combo'):
            from email_sender import EmailTemplate
            template_names = EmailTemplate.get_template_names()
            self.template_combo['values'] = template_names
            if template_names:
                self.template_combo.set(template_names[0])  # Select first template
        
        if hasattr(self, 'billed_invoices_tree'):
            self.refresh_billed_invoices()  # Refresh billed invoices if tab exists
    
    # NEW INVOICE TAB METHODS
    def on_invoice_client_select(self, event):
        """When client is selected, populate projects for that client"""
        client_name = self.invoice_client_combo.get()
        if client_name:
            # Get client ID
            clients = self.client_model.get_all()
            client_id = None
            for client in clients:
                if client[1] == client_name:
                    client_id = client[0]
                    break
            
            if client_id:
                # Load projects for this client
                projects = self.project_model.get_by_client(client_id)
                project_names = [p[2] for p in projects]
                self.invoice_project_combo['values'] = ['All Projects'] + project_names
                self.invoice_project_combo.set('All Projects')
            else:
                self.invoice_project_combo['values'] = []
                self.invoice_project_combo.set('')
    
    def on_invoice_project_select(self, event):
        """Project selection changed - just for filtering later"""
        pass  # Filtering happens when Load button is clicked
    
    def toggle_invoice_date_filter(self):
        """Show or hide date range inputs based on filter selection"""
        if self.invoice_filter_var.get() == "date_range":
            self.date_range_frame.pack(side='left', padx=20)
        else:
            self.date_range_frame.pack_forget()
    
    def refresh_invoice_combos(self):
        """Refresh client and project dropdowns and reload current client's entries"""
        # Save current client selection
        current_client = self.invoice_client_combo.get()
        
        # Refresh client list
        clients = self.client_model.get_all()
        client_names = [c[1] for c in clients]
        self.invoice_client_combo['values'] = client_names
        
        # Restore client selection if it still exists
        if current_client in client_names:
            self.invoice_client_combo.set(current_client)
            # Reload entries for this client
            self.load_invoiceable_entries()
        else:
            self.invoice_project_combo['values'] = []
            self.invoice_project_combo.set('')
            messagebox.showinfo("Refreshed", "Client list updated.")
    
    def load_invoiceable_entries(self):
        """Load unbilled time entries based on selected client/project/date filter"""
        client_name = self.invoice_client_combo.get()
        if not client_name:
            messagebox.showerror("Error", "Please select a client first")
            return
        
        # Get client ID
        clients = self.client_model.get_all()
        client_id = None
        for client in clients:
            if client[1] == client_name:
                client_id = client[0]
                break
        
        if not client_id:
            messagebox.showerror("Error", "Invalid client selected")
            return
        
        # Get project filter
        project_name = self.invoice_project_combo.get()
        project_id = None
        if project_name and project_name != 'All Projects':
            projects = self.project_model.get_by_client(client_id)
            for project in projects:
                if project[2] == project_name:
                    project_id = project[0]
                    break
        
        # Build query based on filter
        conn = self.db.conn  # Use direct connection, not context manager
        cursor = conn.cursor()
        
        if self.invoice_filter_var.get() == "all_uninvoiced":
            # All unbilled entries for client
            if project_id:
                cursor.execute('''
                    SELECT te.id, te.start_time, te.description, te.duration_minutes,
                           p.name as project_name, t.name as task_name
                    FROM time_entries te
                    JOIN tasks t ON te.task_id = t.id
                    JOIN projects p ON te.project_id = p.id
                    WHERE p.client_id = ? AND p.id = ? AND (te.is_billed = 0 OR te.is_billed IS NULL)
                    ORDER BY te.start_time DESC
                ''', (client_id, project_id))
            else:
                cursor.execute('''
                    SELECT te.id, te.start_time, te.description, te.duration_minutes,
                           p.name as project_name, t.name as task_name
                    FROM time_entries te
                    JOIN tasks t ON te.task_id = t.id
                    JOIN projects p ON te.project_id = p.id
                    WHERE p.client_id = ? AND (te.is_billed = 0 OR te.is_billed IS NULL)
                    ORDER BY te.start_time DESC
                ''', (client_id,))
        else:
            # Date range filter
            try:
                start_date = datetime.strptime(self.invoice_start_date.get(), "%m/%d/%y")
                end_date = datetime.strptime(self.invoice_end_date.get(), "%m/%d/%y")
            except ValueError:
                messagebox.showerror("Error", "Invalid date format. Use MM/DD/YY")
                return
            
            if project_id:
                cursor.execute('''
                    SELECT te.id, te.start_time, te.description, te.duration_minutes,
                           p.name as project_name, t.name as task_name
                    FROM time_entries te
                    JOIN tasks t ON te.task_id = t.id
                    JOIN projects p ON te.project_id = p.id
                    WHERE p.client_id = ? AND p.id = ? 
                          AND DATE(te.start_time) BETWEEN ? AND ?
                          AND (te.is_billed = 0 OR te.is_billed IS NULL)
                    ORDER BY te.start_time DESC
                ''', (client_id, project_id, start_date.strftime("%Y-%m-%d"), end_date.strftime("%Y-%m-%d")))
            else:
                cursor.execute('''
                    SELECT te.id, te.start_time, te.description, te.duration_minutes,
                           p.name as project_name, t.name as task_name
                    FROM time_entries te
                    JOIN tasks t ON te.task_id = t.id
                    JOIN projects p ON te.project_id = p.id
                    WHERE p.client_id = ? 
                          AND DATE(te.start_time) BETWEEN ? AND ?
                          AND (te.is_billed = 0 OR te.is_billed IS NULL)
                    ORDER BY te.start_time DESC
                ''', (client_id, start_date.strftime("%Y-%m-%d"), end_date.strftime("%Y-%m-%d")))
        
        entries = cursor.fetchall()
        
        # Clear existing entries
        for item in self.invoice_entries_tree.get_children():
            self.invoice_entries_tree.delete(item)
        
        # Group by Project -> Task for better organization
        project_groups = {}  # {project_name: {task_name: [entries]}}
        
        for entry in entries:
            entry_id, start_time, description, duration_minutes, project_name, task_name = entry
            
            if project_name not in project_groups:
                project_groups[project_name] = {}
            if task_name not in project_groups[project_name]:
                project_groups[project_name][task_name] = []
            
            project_groups[project_name][task_name].append(entry)
        
        # Populate tree with hierarchy
        total_hours = 0
        total_entries = 0
        
        for project_name in sorted(project_groups.keys()):
            # Calculate project totals
            project_hours = 0
            project_entry_count = 0
            for task_name in project_groups[project_name]:
                for entry in project_groups[project_name][task_name]:
                    project_hours += (entry[3] or 0) / 60.0
                    project_entry_count += 1
            
            # Insert project header
            project_node = self.invoice_entries_tree.insert('', 'end',
                text=f'📁 {project_name}',
                values=('', '', '', f'{project_hours:.2f} hrs', f'{project_entry_count} entries'),
                tags=('project', 'project_row'))
            
            # Add tasks under project
            for task_name in sorted(project_groups[project_name].keys()):
                task_entries = project_groups[project_name][task_name]
                
                # Calculate task total
                task_hours = sum((e[3] or 0) / 60.0 for e in task_entries)
                
                # Insert task header
                task_node = self.invoice_entries_tree.insert(project_node, 'end',
                    text=f'  📋 {task_name}',
                    values=('', '', '', f'{task_hours:.2f} hrs', f'{len(task_entries)} entries'),
                    tags=('task', 'task_row'))
                
                # Add individual entries under task
                for entry in task_entries:
                    entry_id, start_time, description, duration_minutes, _, _ = entry
                    
                    # Format date
                    try:
                        dt = datetime.fromisoformat(start_time)
                        date_display = dt.strftime("%m/%d/%y %I:%M %p")
                    except:
                        date_display = start_time[:10]
                    
                    hours = (duration_minutes or 0) / 60.0
                    total_hours += hours
                    total_entries += 1
                    
                    # Insert entry (selectable)
                    self.invoice_entries_tree.insert(task_node, 'end',
                        text='    ⏱️',
                        values=(date_display, '', '', f'{hours:.2f} hrs', description or ''),
                        tags=(f'entry_id_{entry_id}', 'entry', 'entry_row'))
        
        # Save expansion state before clearing (only if tree exists and has items)
        expanded_items = set()
        if hasattr(self, 'invoice_entries_tree'):
            try:
                expanded_items = self.save_tree_state(self.invoice_entries_tree)
            except:
                pass
        
        # Configure tag styles
        self.invoice_entries_tree.tag_configure('project', font=('Arial', 10, 'bold'))
        self.invoice_entries_tree.tag_configure('task', font=('Arial', 9, 'bold'))
        self.invoice_entries_tree.tag_configure('entry', font=('Arial', 9))
        
        # Configure group heading colors for invoice tree
        if 'group_heading' in self.colors:
            self.invoice_entries_tree.tag_configure('project_row', 
                background=self.colors['group_heading'], 
                foreground=self.colors.get('group_text', 'white'),
                font=('Arial', 10, 'bold'))
            
            self.invoice_entries_tree.tag_configure('task_row', 
                background=self.colors['group_heading'], 
                foreground=self.colors.get('group_text', 'white'),
                font=('Arial', 9, 'bold'))
        else:
            # Fallback for themes without group_heading color
            self.invoice_entries_tree.tag_configure('project_row', 
                background='#e8f4f8', 
                foreground='#13100f',
                font=('Arial', 10, 'bold'))
            
            self.invoice_entries_tree.tag_configure('task_row', 
                background='#e8f4f8', 
                foreground='#13100f',
                font=('Arial', 9, 'bold'))
        
        # Configure entry rows (individual time entries) - always white with dark text
        self.invoice_entries_tree.tag_configure('entry_row', 
            background='white', 
            foreground=self.colors['text'],
            font=('Arial', 9))
        
        # Restore expansion state (expand all by default)
        self.restore_tree_state(self.invoice_entries_tree, expanded_items, expand_all=True)
        
        # Update summary
        self.invoice_summary_label.config(
            text=f"📊 {total_entries} unbilled entries found | Total: {total_hours:.2f} hours")
    
    def select_all_invoice_entries(self):
        """Select all ACTUAL entries (not project/task headers) in the invoice tree"""
        # First, expand all nodes so we can select everything
        def expand_all(parent):
            for item in self.invoice_entries_tree.get_children(parent):
                self.invoice_entries_tree.item(item, open=True)
                expand_all(item)
        
        expand_all('')
        
        # Now select all actual entries
        def select_entries_recursive(parent):
            for item in self.invoice_entries_tree.get_children(parent):
                tags = self.invoice_entries_tree.item(item)['tags']
                # Only select items that have entry_id tag (actual time entries)
                if any(tag.startswith('entry_id_') for tag in tags):
                    self.invoice_entries_tree.selection_add(item)
                # Recurse into children
                select_entries_recursive(item)
        
        # Start from root
        select_entries_recursive('')
    
    def deselect_all_invoice_entries(self):
        """Deselect all entries in the invoice tree"""
        # Get all currently selected items and deselect them
        selected = self.invoice_entries_tree.selection()
        if selected:
            self.invoice_entries_tree.selection_remove(*selected)
    
    def edit_invoice_entry(self):
        """Edit a selected time entry from the Invoice Tab"""
        selection = self.invoice_entries_tree.selection()
        if not selection:
            messagebox.showerror("Error", "Please select ONE time entry to edit")
            return
        
        if len(selection) > 1:
            messagebox.showwarning("Multiple Selection", 
                "Please select only ONE entry at a time to edit.\n\n" +
                f"You have selected {len(selection)} entries.")
            return
        
        # Get the selected item
        item = selection[0]
        tags = self.invoice_entries_tree.item(item)['tags']
        
        # Only allow editing of actual entries, not group nodes
        if 'entry' not in tags:
            messagebox.showerror("Error", "Please select an individual time entry (not a project/task group)")
            return
        
        # Extract entry ID from tags
        entry_id = None
        for tag in tags:
            if tag.startswith('entry_id_'):
                entry_id = int(tag.replace('entry_id_', ''))
                break
        
        if not entry_id:
            messagebox.showerror("Error", "Could not find entry ID")
            return
        
        # Open the edit dialog and wait for it to close
        self.open_edit_time_entry_dialog_and_refresh(entry_id)
    
    def preview_invoice(self):
        """Generate and show invoice preview dialog"""
        selection = self.invoice_entries_tree.selection()
        if not selection:
            messagebox.showerror("Error", "Please select at least one time entry to invoice")
            return
        
        # Get selected entry IDs
        entry_ids = []
        for item in selection:
            tags = self.invoice_entries_tree.item(item)['tags']
            for tag in tags:
                if tag.startswith('entry_id_'):
                    entry_ids.append(int(tag.replace('entry_id_', '')))
                    break
        
        if not entry_ids:
            messagebox.showerror("Error", "Could not find selected entries")
            return
        
        # Get client info
        client_name = self.invoice_client_combo.get()
        clients = self.client_model.get_all()
        client_id = None
        for client in clients:
            if client[1] == client_name:
                client_id = client[0]
                break
        
        # Generate invoice data from selected entries
        self.show_invoice_preview_dialog(client_id, client_name, entry_ids)
    
    def show_invoice_preview_dialog(self, client_id, client_name, entry_ids):
        """Show invoice preview dialog with EDIT and CREATE buttons"""
        import sqlite3
        
        # Create dialog
        preview_dialog = tk.Toplevel(self.root)
        preview_dialog.title(f"Invoice Preview - {client_name}")
        self.center_dialog(preview_dialog, 800, 600)
        
        # Get entry details for invoice
        conn = self.db.conn
        conn.row_factory = sqlite3.Row
        cursor = conn.cursor()
        
        placeholders = ','.join(['?' for _ in entry_ids])
        cursor.execute(f'''
            SELECT te.id as entry_id, te.start_time, te.duration_minutes, te.description,
                   p.name as project_name, p.hourly_rate as project_rate, p.is_lump_sum as project_lump_sum,
                   p.lump_sum_amount as project_lump_amount,
                   t.name as task_name, t.hourly_rate as task_rate, t.is_lump_sum as task_lump_sum,
                   t.lump_sum_amount as task_lump_amount
            FROM time_entries te
            JOIN tasks t ON te.task_id = t.id
            JOIN projects p ON te.project_id = p.id
            WHERE te.id IN ({placeholders})
        ''', entry_ids)
        
        entries = cursor.fetchall()
        conn.row_factory = None
        
        # Group by project -> task with subtotals
        project_groups = {}  # {project_name: {tasks: {task_name: data}, subtotal: 0}}
        
        for entry in entries:
            project_name = entry['project_name']
            task_name = entry['task_name']
            
            if project_name not in project_groups:
                project_groups[project_name] = {'tasks': {}, 'subtotal': 0}
            
            if task_name not in project_groups[project_name]['tasks']:
                project_groups[project_name]['tasks'][task_name] = {
                    'minutes': 0,
                    'rate': entry['task_rate'] or entry['project_rate'],
                    'is_lump_sum': entry['task_lump_sum'] or entry['project_lump_sum'],
                    'lump_sum_amount': entry['task_lump_amount'] or entry['project_lump_amount']
                }
            
            project_groups[project_name]['tasks'][task_name]['minutes'] += entry['duration_minutes'] or 0
        
        # Build invoice items with project/task hierarchy
        invoice_items = []
        total_amount = 0
        total_hours = 0  # Track total hours across all tasks
        
        for project_name, project_data in project_groups.items():
            # Add project header
            invoice_items.append({
                'description': f'**{project_name}**',
                'quantity': '',
                'rate': '',
                'amount': '',
                'is_header': True
            })
            
            project_subtotal = 0
            
            for task_name, task_data in project_data['tasks'].items():
                hours = task_data['minutes'] / 60.0
                total_hours += hours  # Add to running total
                
                if task_data['is_lump_sum']:
                    amount = task_data['lump_sum_amount']
                    invoice_items.append({
                        'description': f'  • {task_name}',
                        'quantity': '1',
                        'rate': f"${amount:.2f}",
                        'amount': amount,
                        'is_task': True
                    })
                else:
                    amount = hours * task_data['rate']
                    invoice_items.append({
                        'description': f'  • {task_name}',
                        'quantity': f"{hours:.2f} hrs",
                        'rate': f"${task_data['rate']:.2f}/hr",
                        'amount': amount,
                        'is_task': True
                    })
                
                project_subtotal += amount
            
            # Add project subtotal
            invoice_items.append({
                'description': f'  {project_name} Subtotal',
                'quantity': '',
                'rate': '',
                'amount': project_subtotal,
                'is_subtotal': True
            })
            
            total_amount += project_subtotal
        
        # Get date range from entries
        start_dates = [datetime.fromisoformat(e['start_time']) for e in entries]
        start_date = min(start_dates) if start_dates else datetime.now()
        end_date = max(start_dates) if start_dates else datetime.now()
        
        # Store for later use
        self.current_invoice_data = {
            'client_id': client_id,
            'client_name': client_name,
            'entry_ids': entry_ids,
            'items': invoice_items,
            'total': total_amount,
            'start_date': start_date,
            'end_date': end_date
        }
        
        # Header
        header_frame = ttk.Frame(preview_dialog)
        header_frame.pack(fill='x', padx=20, pady=20)
        
        ttk.Label(header_frame, text=f"INVOICE PREVIEW", 
                 font=('Arial', 16, 'bold')).pack()
        ttk.Label(header_frame, text=f"Client: {client_name}", 
                 font=('Arial', 12)).pack(pady=5)
        ttk.Label(header_frame, text=f"Date: {datetime.now().strftime('%B %d, %Y')}", 
                 font=('Arial', 10)).pack()
        
        # Items tree
        items_frame = ttk.LabelFrame(preview_dialog, text="Invoice Items")
        items_frame.pack(fill='both', expand=True, padx=20, pady=10)
        
        items_tree = ttk.Treeview(items_frame, 
            columns=('Description', 'Quantity', 'Rate', 'Amount'),
            show='headings')
        
        items_tree.heading('Description', text='Description')
        items_tree.heading('Quantity', text='Quantity')
        items_tree.heading('Rate', text='Rate')
        items_tree.heading('Amount', text='Amount')
        
        items_tree.column('Description', width=350)
        items_tree.column('Quantity', width=100)
        items_tree.column('Rate', width=100)
        items_tree.column('Amount', width=100)
        
        for item in invoice_items:
            if item.get('is_header'):
                # Bold project header
                items_tree.insert('', 'end', values=(
                    item['description'].replace('**', ''),
                    '', '', ''
                ), tags=('header',))
            elif item.get('is_subtotal'):
                # Project subtotal row
                items_tree.insert('', 'end', values=(
                    item['description'],
                    '', '',
                    f"${item['amount']:.2f}"
                ), tags=('subtotal',))
            else:
                # Regular task row
                amount_display = f"${item['amount']:.2f}" if isinstance(item['amount'], (int, float)) else ''
                items_tree.insert('', 'end', values=(
                    item['description'],
                    item['quantity'],
                    item['rate'],
                    amount_display
                ))
        
        # Configure tags for styling
        items_tree.tag_configure('header', font=('Arial', 10, 'bold'), background='#e8f4f8')
        items_tree.tag_configure('subtotal', font=('Arial', 9, 'bold'), background='#f0f0f0')
        
        items_tree.pack(fill='both', expand=True, padx=10, pady=10)
        
        # Total
        total_frame = ttk.Frame(preview_dialog)
        total_frame.pack(fill='x', padx=20, pady=10)
        
        # Total hours label (left side)
        ttk.Label(total_frame, text=f"Total Hours: {total_hours:.2f} hrs", 
                 font=('Arial', 12, 'bold'),
                 foreground=self.colors['text_secondary']).pack(side='left')
        
        # Total amount label (right side)
        ttk.Label(total_frame, text=f"TOTAL: ${total_amount:.2f}", 
                 font=('Arial', 14, 'bold')).pack(side='right')
        
        # Buttons
        button_frame = ttk.Frame(preview_dialog)
        button_frame.pack(fill='x', padx=20, pady=20)
        
        def create_invoice():
            """Mark entries as billed and generate PDF"""
            if messagebox.askyesno("Confirm", 
                f"Create invoice for ${total_amount:.2f}?\n\n" +
                "This will mark all selected time entries as BILLED."):
                
                # Generate invoice number
                invoice_date = datetime.now()
                invoice_number = f"INV-{invoice_date.strftime('%Y%m%d-%H%M%S')}"
                
                # Ask for save location
                filename = filedialog.asksaveasfilename(
                    defaultextension=".pdf",
                    filetypes=[("PDF files", "*.pdf")],
                    initialfile=f"Invoice_{client_name.replace(' ', '_')}_{invoice_number}.pdf"
                )
                
                if filename:
                    try:
                        # Generate PDF
                        from invoice_generator import InvoiceGenerator
                        generator = InvoiceGenerator(self.db)
                        
                        print(f"[DEBUG] Generating PDF with data: {self.current_invoice_data.keys()}")
                        print(f"[DEBUG] Start date: {self.current_invoice_data['start_date']}")
                        print(f"[DEBUG] End date: {self.current_invoice_data['end_date']}")
                        print(f"[DEBUG] Total items: {len(self.current_invoice_data['items'])}")
                        
                        generator.generate_pdf(self.current_invoice_data, filename, invoice_number)
                        
                        print(f"[DEBUG] PDF generated successfully at: {filename}")
                        
                        # Mark entries as billed
                        conn = self.db.conn  # Use direct connection
                        cursor = conn.cursor()
                        placeholders = ','.join(['?' for _ in entry_ids])
                        cursor.execute(f'''
                            UPDATE time_entries 
                            SET is_billed = 1, 
                                invoice_number = ?,
                                billing_date = ?
                            WHERE id IN ({placeholders})
                        ''', [invoice_number, invoice_date.strftime("%Y-%m-%d")] + entry_ids)
                        conn.commit()
                        
                        print(f"[DEBUG] Marked {len(entry_ids)} entries as billed")
                        
                        # Refresh displays
                        self.refresh_time_entries()
                        self.load_invoiceable_entries()
                        
                        preview_dialog.destroy()
                        
                        messagebox.showinfo("Success", 
                            f"Invoice created successfully!\n\n" +
                            f"File: {filename}\n" +
                            f"Invoice #: {invoice_number}\n" +
                            f"Total: ${total_amount:.2f}\n\n" +
                            f"{len(entry_ids)} time entries marked as billed.")
                    
                    except Exception as e:
                        import traceback
                        error_details = traceback.format_exc()
                        print(f"[ERROR] Invoice generation failed:")
                        print(error_details)
                        messagebox.showerror("Error", 
                            f"Failed to create invoice:\n\n{str(e)}\n\n" +
                            f"Check console for details.")
        
        def edit_time_entries_from_preview():
            """Allow editing time entries from the invoice preview"""
            # Create a selection window showing all time entries in this invoice
            edit_window = tk.Toplevel(preview_dialog)
            edit_window.title("Edit Time Entries")
            self.center_dialog(edit_window, 700, 500)
            
            # Header
            ttk.Label(edit_window, text="Select a time entry to edit:", 
                     font=('Arial', 12, 'bold')).pack(padx=20, pady=10)
            
            # Tree showing all entries
            tree_frame = ttk.Frame(edit_window)
            tree_frame.pack(fill='both', expand=True, padx=20, pady=10)
            
            edit_tree = ttk.Treeview(tree_frame,
                columns=('Date', 'Project', 'Task', 'Hours', 'Description'),
                show='headings',
                selectmode='browse')
            
            edit_tree.heading('Date', text='Date')
            edit_tree.heading('Project', text='Project')
            edit_tree.heading('Task', text='Task')
            edit_tree.heading('Hours', text='Hours')
            edit_tree.heading('Description', text='Description')
            
            edit_tree.column('Date', width=120)
            edit_tree.column('Project', width=120)
            edit_tree.column('Task', width=120)
            edit_tree.column('Hours', width=80)
            edit_tree.column('Description', width=200)
            
            # Scrollbar for tree
            scrollbar = ttk.Scrollbar(tree_frame, orient='vertical', command=edit_tree.yview)
            edit_tree.configure(yscrollcommand=scrollbar.set)
            
            edit_tree.pack(side='left', fill='both', expand=True)
            scrollbar.pack(side='right', fill='y')
            
            # Load time entry details
            conn = self.db.conn
            conn.row_factory = sqlite3.Row
            cursor = conn.cursor()
            
            placeholders = ','.join(['?' for _ in entry_ids])
            cursor.execute(f'''
                SELECT te.id as entry_id, te.start_time, te.end_time, te.duration_minutes, 
                       te.description, p.name as project_name, t.name as task_name
                FROM time_entries te
                JOIN tasks t ON te.task_id = t.id
                JOIN projects p ON te.project_id = p.id
                WHERE te.id IN ({placeholders})
                ORDER BY te.start_time
            ''', entry_ids)
            
            time_entries = cursor.fetchall()
            conn.row_factory = None
            
            # Populate tree
            for entry in time_entries:
                try:
                    dt = datetime.fromisoformat(entry['start_time'])
                    date_display = dt.strftime("%m/%d/%y %I:%M %p")
                except:
                    date_display = entry['start_time'][:10]
                
                hours = (entry['duration_minutes'] or 0) / 60.0
                
                edit_tree.insert('', 'end',
                    values=(
                        date_display,
                        entry['project_name'],
                        entry['task_name'],
                        f"{hours:.2f}",
                        entry['description'] or ''
                    ),
                    tags=(f"entry_id_{entry['entry_id']}",))
            
            # Button frame
            btn_frame = ttk.Frame(edit_window)
            btn_frame.pack(fill='x', padx=20, pady=20)
            
            def edit_selected_entry():
                """Edit the selected time entry"""
                selection = edit_tree.selection()
                if not selection:
                    messagebox.showerror("Error", "Please select a time entry to edit")
                    return
                
                # Extract entry ID from tags
                tags = edit_tree.item(selection[0])['tags']
                entry_id = None
                for tag in tags:
                    if tag.startswith('entry_id_'):
                        entry_id = int(tag.replace('entry_id_', ''))
                        break
                
                if not entry_id:
                    messagebox.showerror("Error", "Could not find entry ID")
                    return
                
                # Close the edit selection window
                edit_window.destroy()
                
                # Open the edit dialog (reuse existing method)
                self.open_edit_time_entry_dialog(entry_id)
                
                # Note: User must manually refresh using "Refresh & Close" button after editing
                # This prevents the confusing auto-refresh before user has made changes
            
            def refresh_invoice_preview():
                """Refresh the invoice preview with updated data"""
                edit_window.destroy()
                preview_dialog.destroy()
                # Reload the invoice entries to show updated data
                self.load_invoiceable_entries()
                messagebox.showinfo("Refreshed", 
                    "Invoice data refreshed. Select entries and preview again to see changes.")
            
            ttk.Button(btn_frame, text="✏️ Edit Selected Entry", 
                      command=edit_selected_entry).pack(side='left', padx=5)
            ttk.Button(btn_frame, text="🔄 Refresh & Close", 
                      command=refresh_invoice_preview).pack(side='left', padx=5)
            ttk.Button(btn_frame, text="Cancel", 
                      command=edit_window.destroy).pack(side='right', padx=5)
        
        def email_invoice():
            """Email the invoice to the client"""
            # First, check if email is configured
            email_settings = self.db.get_email_settings()
            if not email_settings:
                if messagebox.askyesno("Email Not Configured",
                    "Email settings haven't been configured yet.\n\n" +
                    "Would you like to set them up now?"):
                    # Switch to Email Settings tab
                    self.notebook.select(6)  # Email Settings tab index
                    preview_dialog.destroy()
                return
            
            # Get client email
            client = self.client_model.get_by_id(client_id)
            if not client or not client[3]:  # client[3] is email
                messagebox.showerror("No Email Address",
                    f"Client '{client_name}' doesn't have an email address on file.\n\n" +
                    "Please add their email in the Clients tab first.")
                return
            
            client_email = client[3]
            
            # Show email send dialog
            self.show_email_invoice_dialog(preview_dialog, client_name, client_email, 
                                           client_id, entry_ids, invoice_items, total_amount,
                                           start_date, end_date)
        
        ttk.Button(button_frame, text="✏️ Edit Entries", 
                  command=edit_time_entries_from_preview).pack(side='left', padx=5)
        ttk.Button(button_frame, text="📧 Email Invoice", 
                  command=email_invoice).pack(side='right', padx=5)
        ttk.Button(button_frame, text="✅ CREATE INVOICE", 
                  command=create_invoice,
                  style='Accent.TButton').pack(side='right', padx=5)
        ttk.Button(button_frame, text="Cancel", 
                  command=preview_dialog.destroy).pack(side='right', padx=5)
    
    def show_email_invoice_dialog(self, parent_dialog, client_name, client_email, 
                                   client_id, entry_ids, invoice_items, total_amount,
                                   start_date, end_date):
        """Show dialog to send invoice via email"""
        import tempfile
        from email_sender import EmailSender, EmailTemplate
        
        # Create email dialog
        email_dialog = tk.Toplevel(parent_dialog)
        email_dialog.title("Email Invoice")
        email_dialog.geometry("600x700")
        email_dialog.transient(parent_dialog)
        email_dialog.grab_set()
        
        # Header
        header_frame = ttk.Frame(email_dialog)
        header_frame.pack(fill='x', padx=20, pady=20)
        
        ttk.Label(header_frame, text="📧 Email Invoice", 
                 font=('Arial', 16, 'bold')).pack()
        ttk.Label(header_frame, text=f"To: {client_name} ({client_email})", 
                 font=('Arial', 10)).pack(pady=5)
        
        # Form
        form_frame = ttk.LabelFrame(email_dialog, text="Email Details")
        form_frame.pack(fill='both', expand=True, padx=20, pady=10)
        
        # Template selection
        ttk.Label(form_frame, text="Template:").grid(row=0, column=0, sticky='w', padx=10, pady=5)
        template_var = tk.StringVar(value="Professional")
        template_combo = ttk.Combobox(form_frame, textvariable=template_var, 
                                     values=EmailTemplate.get_template_names(), state='readonly')
        template_combo.grid(row=0, column=1, sticky='ew', padx=10, pady=5)
        
        # Subject
        ttk.Label(form_frame, text="Subject:").grid(row=1, column=0, sticky='w', padx=10, pady=5)
        subject_entry = ttk.Entry(form_frame, width=50)
        subject_entry.grid(row=1, column=1, sticky='ew', padx=10, pady=5)
        
        # CC (optional)
        ttk.Label(form_frame, text="CC (optional):").grid(row=2, column=0, sticky='w', padx=10, pady=5)
        cc_entry = ttk.Entry(form_frame, width=50)
        cc_entry.grid(row=2, column=1, sticky='ew', padx=10, pady=5)
        ttk.Label(form_frame, text="Separate multiple emails with commas", 
                 font=('Arial', 8), foreground='gray').grid(row=3, column=1, sticky='w', padx=10)
        
        # Message body
        ttk.Label(form_frame, text="Message:").grid(row=4, column=0, sticky='nw', padx=10, pady=5)
        message_text = tk.Text(form_frame, height=15, wrap='word')
        message_text.grid(row=4, column=1, sticky='ew', padx=10, pady=5)
        
        # Scrollbar for message
        message_scroll = ttk.Scrollbar(form_frame, orient='vertical', command=message_text.yview)
        message_text.configure(yscrollcommand=message_scroll.set)
        message_scroll.grid(row=4, column=2, sticky='ns', pady=5)
        
        form_frame.columnconfigure(1, weight=1)
        
        # Function to update template preview
        def update_template(*args):
            template_name = template_var.get()
            template = EmailTemplate.get_template(template_name)
            if template:
                # Get company info
                company = self.company_model.get()
                
                # Prepare variables for template
                invoice_number = f"INV-{datetime.now().strftime('%Y%m%d-%H%M%S')}"
                date_range = f"{start_date.strftime('%m/%d/%y')} - {end_date.strftime('%m/%d/%y')}"
                
                variables = {
                    'client_name': client_name,
                    'client_email': client_email,
                    'invoice_number': invoice_number,
                    'invoice_date': datetime.now().strftime('%B %d, %Y'),
                    'invoice_total': f"${total_amount:.2f}",
                    'payment_terms': company[7] if company and len(company) > 7 else 'Net 30',
                    'due_date': (datetime.now() + timedelta(days=30)).strftime('%B %d, %Y'),
                    'date_range': date_range,
                    'company_name': company[1] if company else 'Your Company',
                    'company_email': company[4] if company and len(company) > 4 else '',
                    'company_phone': company[3] if company and len(company) > 3 else '',
                    'company_website': company[6] if company and len(company) > 6 else ''
                }
                
                # Render template
                subject = EmailTemplate.render_template(template['subject'], variables)
                body = EmailTemplate.render_template(template['body'], variables)
                
                # Update fields
                subject_entry.delete(0, tk.END)
                subject_entry.insert(0, subject)
                
                message_text.delete('1.0', tk.END)
                message_text.insert('1.0', body)
        
        # Bind template change
        template_combo.bind('<<ComboboxSelected>>', update_template)
        
        # Load initial template
        update_template()
        
        # Buttons
        button_frame = ttk.Frame(email_dialog)
        button_frame.pack(fill='x', padx=20, pady=20)
        
        def send_email():
            """Send the email with invoice attached"""
            try:
                # Get email settings
                email_settings = self.db.get_email_settings()
                if not email_settings:
                    messagebox.showerror("Error", "Email settings not found")
                    return
                
                smtp_server = email_settings[1]
                smtp_port = email_settings[2]
                email_address = email_settings[3]
                email_password = email_settings[4]
                from_name = email_settings[5] if len(email_settings) > 5 else None
                
                # Initialize email sender
                sender = EmailSender(smtp_server, smtp_port, email_address, email_password)
                
                # Generate invoice PDF in temp location
                invoice_number = f"INV-{datetime.now().strftime('%Y%m%d-%H%M%S')}"
                temp_dir = tempfile.gettempdir()
                pdf_path = os.path.join(temp_dir, f"{invoice_number}.pdf")
                
                # Create invoice data
                invoice_data = {
                    'client_id': client_id,
                    'client_name': client_name,
                    'entry_ids': entry_ids,
                    'items': invoice_items,
                    'total': total_amount,
                    'start_date': start_date,
                    'end_date': end_date
                }
                
                # Generate PDF
                from invoice_generator import InvoiceGenerator
                generator = InvoiceGenerator(self.db)
                generator.generate_pdf(invoice_data, pdf_path, invoice_number)
                
                # Get email details
                subject = subject_entry.get().strip()
                body_html = message_text.get('1.0', tk.END).strip()
                
                # Parse CC addresses
                cc_addresses = None
                cc_text = cc_entry.get().strip()
                if cc_text:
                    cc_addresses = [email.strip() for email in cc_text.split(',')]
                
                # Send email
                success, message = sender.send_email(
                    to_address=client_email,
                    subject=subject,
                    body_html=body_html,
                    attachment_path=pdf_path,
                    cc_addresses=cc_addresses,
                    from_name=from_name
                )
                
                # Clean up temp PDF
                try:
                    os.remove(pdf_path)
                except:
                    pass
                
                if success:
                    # Ask if they want to mark as billed
                    if messagebox.askyesno("Email Sent!",
                        f"Invoice emailed successfully to {client_email}!\n\n" +
                        "Would you like to mark these time entries as BILLED now?"):
                        
                        # Mark entries as billed
                        invoice_date = datetime.now()
                        conn = self.db.conn
                        cursor = conn.cursor()
                        placeholders = ','.join(['?' for _ in entry_ids])
                        cursor.execute(f'''
                            UPDATE time_entries 
                            SET is_billed = 1, 
                                invoice_number = ?,
                                billing_date = ?
                            WHERE id IN ({placeholders})
                        ''', [invoice_number, invoice_date.strftime("%Y-%m-%d")] + entry_ids)
                        conn.commit()
                        
                        # Refresh displays
                        self.refresh_time_entries()
                        self.load_invoiceable_entries()
                        
                        messagebox.showinfo("Success",
                            f"Invoice #{invoice_number} sent and {len(entry_ids)} entries marked as billed!")
                    
                    # Close dialogs
                    email_dialog.destroy()
                    parent_dialog.destroy()
                else:
                    messagebox.showerror("Send Failed", message)
                    
            except Exception as e:
                import traceback
                traceback.print_exc()
                messagebox.showerror("Error", f"Failed to send email:\n\n{str(e)}")
        
        ttk.Button(button_frame, text="📧 Send Invoice", 
                  command=send_email,
                  style='Accent.TButton').pack(side='right', padx=5)
        ttk.Button(button_frame, text="Cancel", 
                  command=email_dialog.destroy).pack(side='right', padx=5)
    
